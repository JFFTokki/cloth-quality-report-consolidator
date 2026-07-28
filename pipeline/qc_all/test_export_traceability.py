import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from source_relationship import source_relationship_id


ROOT = Path(__file__).resolve().parents[2]
WRAPPER = ROOT / "quality-report-consolidator" / "scripts" / "validate_and_export.mjs"
EXPORTER = ROOT / "quality-report-consolidator" / "scripts" / "export_leader_workbook.mjs"


def relation(row_number):
    return source_relationship_id(
        workbook="source.xlsx",
        sheet="Sheet1",
        row=row_number,
        cell=f"J{row_number}",
        url="https://example.test/shared.pdf",
        sku="200000000001",
        color="黑色",
        sample_type="成衣",
    )


def main():
    source_index = []
    detail_rows = []
    for source_row in (2, 3):
        relationship_id = relation(source_row)
        source_index.append({
            "source_relationship_id": relationship_id,
            "source_sheet": "Sheet1",
            "source_row": source_row,
            "source_cell": f"J{source_row}",
            "url": "https://example.test/shared.pdf",
            "sku": "200000000001",
            "color": "黑色",
            "sample_type": "成衣",
            "processing_status": "已解析",
            "order_no": f"ORDER-{source_row}",
            "overall_result": "合格",
        })
        detail_rows.append({
            "source_relationship_id": relationship_id,
            "source_sheet": "Sheet1",
            "source_row": source_row,
            "source_cell": f"J{source_row}",
            "sku": "200000000001",
            "color": "黑色",
            "sample_type": "成衣",
            "source_order_no": f"ORDER-{source_row}",
            "url": "https://example.test/shared.pdf",
            "raw_item": "pH值",
            "simple_item": "pH值",
            "standard_item": "pH值",
            "status": "自动归并",
            "include_horizontal": True,
            "report_issue_date": "2026-02-31" if source_row == 2 else "2024-02-29",
            "report_issue_date_status": "已识别",
            "report_issue_date_reason": "历史识别结果",
        })

    payload = {
        "source_workbook": "source.xlsx",
        "sample_skus": ["200000000001"],
        "sample_urls": ["https://example.test/shared.pdf"],
        "source_index": source_index,
        "detail_rows": detail_rows,
        "errors": [],
        "report_metadata_by_url": {
            "https://example.test/shared.pdf": {
                "report_issue_date": "",
                "report_issue_date_status": "未发现",
                "report_issue_date_reason": "测试数据未提供签发日期",
                "institution": "测试检验机构",
                "cma_mark": "未发现",
                "cnas_mark": "未发现",
                "cma_evidence": "测试数据未发现CMA标识",
                "cnas_evidence": "测试数据未发现CNAS标识",
            },
        },
    }
    with tempfile.TemporaryDirectory() as temp_dir:
        temp = Path(temp_dir)
        data_path = temp / "report_data.json"
        downloads_path = temp / "download_manifest.json"
        output_path = temp / "workbook.xlsx"
        data_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        downloads_path.write_text(json.dumps([{
            "url": "https://example.test/shared.pdf",
            "status": "downloaded",
        }]), encoding="utf-8")
        environment = os.environ.copy()
        environment["QC_PYTHON"] = sys.executable
        result = subprocess.run(
            ["node", str(WRAPPER), str(data_path), str(downloads_path), str(output_path)],
            cwd=ROOT,
            capture_output=True,
            text=True,
            env=environment,
        )
        assert result.returncode == 3, result.stderr or result.stdout
        review_output_path = output_path.with_name("workbook_external-review-required.xlsx")
        assert not output_path.exists()
        assert review_output_path.exists()
        assert review_output_path.with_suffix(".validation.json").exists()
        assert review_output_path.with_suffix(".export.json").exists()
        workbook = load_workbook(review_output_path, data_only=False)
        sheet = workbook["质检全项目明细"]
        headers = [cell.value for cell in sheet[5]]
        header_index = {header: index for index, header in enumerate(headers)}
        rows = [
            list(row) for row in sheet.iter_rows(min_row=6, values_only=True)
            if row[header_index["记录类型"]] == "检测结果明细"
        ]
        assert [row[header_index["来源表行"]] for row in rows] == [2, 3]
        assert [row[header_index["来源关系ID"]] for row in rows] == [relation(2), relation(3)]
        assert rows[0][header_index["报告签发日期"]] in (None, "")
        assert rows[0][header_index["报告签发日期识别状态"]] == "待复核"
        assert "非法日历日期" in rows[0][header_index["报告签发日期异常原因"]]
        assert rows[1][header_index["报告签发日期"]] is not None
        assert all(row[header_index["规则版本"]] == "rule-unspecified" for row in rows)
        workbook.close()

        legacy_payload = {
            "source_workbook": "source.xlsx",
            "sample_skus": ["200000000002"],
            "sample_urls": ["https://example.test/multicolor.pdf"],
            "source_index": [{
                "source_sheet": "Sheet1",
                "source_row": 8,
                "source_cell": "J8",
                "url": "https://example.test/multicolor.pdf",
                "sku": "200000000002",
                "selected_colors": "黑色，白色",
                "sample_type": "成衣",
            }],
            "detail_rows": [{
                "source_sheet": "Sheet1",
                "source_row": 8,
                "source_cell": "J8",
                "url": "https://example.test/multicolor.pdf",
                "sku": "200000000002",
                "color": color,
                "sample_type": "成衣",
                "raw_item": "pH值",
                "simple_item": "pH值",
                "standard_item": "pH值",
                "status": "自动归并",
                "include_horizontal": True,
            } for color in ("黑色", "白色")],
            "errors": [],
            "report_metadata_by_url": {},
        }
        legacy_data_path = temp / "legacy_report_data.json"
        legacy_output_path = temp / "legacy_workbook.xlsx"
        legacy_data_path.write_text(json.dumps(legacy_payload, ensure_ascii=False), encoding="utf-8")
        legacy_result = subprocess.run(
            ["node", str(EXPORTER), str(legacy_data_path), str(legacy_output_path)],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )
        assert legacy_result.returncode == 0, legacy_result.stderr or legacy_result.stdout
        legacy_workbook = load_workbook(legacy_output_path, data_only=False)
        legacy_sheet = legacy_workbook["质检全项目明细"]
        legacy_headers = [cell.value for cell in legacy_sheet[5]]
        legacy_header_index = {header: index for index, header in enumerate(legacy_headers)}
        legacy_rows = [
            list(row) for row in legacy_sheet.iter_rows(min_row=6, values_only=True)
            if row[legacy_header_index["记录类型"]] == "检测结果明细"
        ]
        relationship_ids = [row[legacy_header_index["来源关系ID"]] for row in legacy_rows]
        assert len(set(relationship_ids)) == 2
        assert all(row[legacy_header_index["规则版本"]] == "rule-unspecified" for row in legacy_rows)
        legacy_workbook.close()

        canonical_id = source_relationship_id(
            workbook="canonical.xlsx",
            sheet="Data",
            row=2,
            cell="H2",
            url="https://example.test/canonical.pdf",
            sku="200000000003",
            color="蓝色",
            sample_type="成衣",
        )
        canonical_payload = {
            "schemaVersion": "quality-report-consolidator/v1",
            "sourceWorkbook": "canonical.xlsx",
            "sourceRelationships": [{
                "sourceRelationshipId": canonical_id,
                "sourceSheet": "Data",
                "sourceRow": 2,
                "sourceCell": "H2",
                "url": "https://example.test/canonical.pdf",
                "sourceProductCode": "200000000003",
                "color": "蓝色",
                "sampleType": "成衣",
                "processStatus": "succeeded",
            }],
            "records": [{
                "recordId": "rec-canonical-1",
                "recordType": "检测结果明细",
                "sourceRelationshipId": canonical_id,
                "sourceSheet": "Data",
                "sourceRow": 2,
                "sourceCell": "H2",
                "url": "https://example.test/canonical.pdf",
                "sourceProductCode": "200000000003",
                "color": "蓝色",
                "sampleType": "成衣",
                "rawItem": "pH值",
                "simplifiedItem": "pH值",
                "parentItem": "pH值",
                "mappingStatus": "自动归并",
                "includeInOverview": True,
                "processStatus": "succeeded",
                "localPath": "/tmp/canonical.pdf",
                "parserVersion": "record-parser-v1",
                "ruleVersion": "record-rule-v1",
            }],
            "reportMetadataByUrl": {},
            "pipelineVersions": {"mapping_rules": "envelope-rule-v1"},
            "classificationDefault": "待分类",
        }
        canonical_data_path = temp / "canonical_report_data.json"
        canonical_output_path = temp / "canonical_workbook.xlsx"
        canonical_data_path.write_text(json.dumps(canonical_payload, ensure_ascii=False), encoding="utf-8")
        canonical_result = subprocess.run(
            ["node", str(EXPORTER), str(canonical_data_path), str(canonical_output_path)],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )
        assert canonical_result.returncode == 0, canonical_result.stderr or canonical_result.stdout
        canonical_workbook = load_workbook(canonical_output_path, data_only=False)
        canonical_sheet = canonical_workbook["质检全项目明细"]
        canonical_headers = [cell.value for cell in canonical_sheet[5]]
        canonical_header_index = {header: index for index, header in enumerate(canonical_headers)}
        canonical_row = next(
            list(row) for row in canonical_sheet.iter_rows(min_row=6, values_only=True)
            if row[canonical_header_index["记录类型"]] == "检测结果明细"
        )
        assert canonical_row[canonical_header_index["记录ID"]] == "rec-canonical-1"
        assert canonical_row[canonical_header_index["处理状态"]] == "succeeded"
        assert canonical_row[canonical_header_index["本地PDF路径"]] == "/tmp/canonical.pdf"
        assert canonical_row[canonical_header_index["解析器版本"]] == "record-parser-v1"
        assert canonical_row[canonical_header_index["规则版本"]] == "record-rule-v1"
        canonical_workbook.close()
        overwrite_result = subprocess.run(
            ["node", str(EXPORTER), str(canonical_data_path), str(canonical_output_path)],
            cwd=ROOT,
            capture_output=True,
            text=True,
        )
        assert overwrite_result.returncode != 0
        assert "Refusing to overwrite existing output" in (overwrite_result.stderr + overwrite_result.stdout)
    print("export traceability and calendar contract ok")


if __name__ == "__main__":
    main()
