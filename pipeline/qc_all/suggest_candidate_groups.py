import json, re
from collections import defaultdict
from pathlib import Path
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

root = Path.cwd()
items = json.loads((root/'pipeline/qc_all/item_candidates.json').read_text(encoding='utf-8'))
item_by_name = {x['simple_item']: x for x in items}

# 明确的语义归并种子：只做建议，不自动写入正式规则。
manual_groups = {
    '耐摩擦色牢度': ['干摩', '干摩擦', '湿摩', '湿摩擦', '干摩擦色牢度', '湿摩擦色牢度', '耐摩擦色牢度'],
    '缝子纰裂程度': ['缝子纰裂', '接缝性能缝子纰裂', '现象判定接缝性能', '缝子纰裂程度', '接缝性能'],
    '沾色': ['沾色', '耐唾液色牢腈纶', '酸汗碱聚酯'],
    '变色': ['变色'],
    '防泼水性能': ['防泼水性能', '拒水性', '防水性能'],
    '防油性能': ['防油性能', '拒油性'],
    '防污性能': ['防污性能', '耐沾污性'],
    '掉纤测试': ['掉纤测试', '掉纤测试森马集团针织服装'],
    '杂质含量': ['杂质', '杂质含量'],
    '陆禽毛含量': ['陆禽毛', '陆禽毛含量'],
    '绒丝+羽丝含量': ['绒丝羽丝', '绒丝羽丝含量'],
    '重金属': ['重金属', '可萃取汞', '镉'],
    '附件锐利性': ['附件锐利尖端', '附件锐利边缘', '金属附件允许轻微掉漆'],
    '附件尺寸': ['附件尺寸'],
    '防风性能': ['防风性能'],
    '透湿性能': ['透湿性能'],
    '勾丝性能': ['勾丝性能'],
    '延伸值': ['横向延伸值', '直向延伸值'],
    '维护方法': ['维护方法', '维护方法无'],
}

noise_keywords = ['没有此项', '岁以下儿童没有此项', '表中表示', '本项判定', '测定低限', '批号款号', '结论', '本机构', '识还应含有产品质量检验', '其他']

assigned = {}
groups = []

def add_group(std, names, reason):
    names = [n for n in dict.fromkeys(names) if n in item_by_name]
    if len(names) < 2:
        return
    gid = f'G{len(groups)+1:03d}'
    total = sum(item_by_name[n]['count'] for n in names)
    for n in names:
        assigned[n] = gid
    groups.append({'gid': gid, 'suggested': std, 'names': names, 'total': total, 'reason': reason})

for std, seeds in manual_groups.items():
    names = []
    for name in item_by_name:
        for seed in seeds:
            if name == seed or seed in name or name in seed:
                names.append(name)
                break
    add_group(std, names, '人工种子词匹配：名称包含相同核心检测语义')

# 针对禁用偶氮染料拆出来的芳香胺物质，建议归入一个大类，但保留各物质名称供判断。
amine_tokens = ['苯胺', '联苯胺', '甲苯胺', '萘胺', '氨基偶氮', '二氨基', '二甲基', '二甲氧基', '亚甲基二', '对氯苯胺']
amines = [n for n in item_by_name if n not in assigned and any(t in n for t in amine_tokens)]
add_group('可分解致癌芳香胺染料-具体芳香胺物质', amines, '疑似同属禁用偶氮/芳香胺具体物质明细，建议作为同一大类下子项')

# 根据简化规范名做相似归并：去掉常见修饰词后完全相同或高度相似。
def norm(name):
    s = name
    for token in ['性能', '测试', '试验', '含量', '程度', '项目', '采购内控标准', '森马集团', '针织服装', '客户要求']:
        s = s.replace(token, '')
    s = s.replace('乾', '干').replace('溼', '湿')
    return s

remaining = [n for n in item_by_name if n not in assigned and not any(k in n for k in noise_keywords)]
buckets = defaultdict(list)
for n in remaining:
    key = norm(n)
    if len(key) >= 2:
        buckets[key].append(n)
for key, names in buckets.items():
    add_group(min(names, key=len), names, '去除“性能/测试/含量/程度”等修饰词后名称一致')

remaining = [n for n in item_by_name if n not in assigned and not any(k in n for k in noise_keywords)]
used = set()
for i, a in enumerate(remaining):
    if a in used:
        continue
    cluster = [a]
    na = norm(a)
    for b in remaining[i+1:]:
        if b in used:
            continue
        nb = norm(b)
        if min(len(na), len(nb)) < 3:
            continue
        ratio = SequenceMatcher(None, na, nb).ratio()
        contain = na in nb or nb in na
        if ratio >= 0.82 or (contain and min(len(na), len(nb)) >= 4):
            cluster.append(b)
    if len(cluster) >= 2:
        for n in cluster:
            used.add(n)
        add_group(min(cluster, key=len), cluster, '文本相似度较高或核心词互相包含，需人工复核')

# 疑似非检测项单独列出，方便人工判定是否排除。
noise_names = [n for n in item_by_name if any(k in n for k in noise_keywords)]
if noise_names:
    gid = f'G{len(groups)+1:03d}'
    groups.append({'gid': gid, 'suggested': '疑似非检测项-建议排除', 'names': noise_names, 'total': sum(item_by_name[n]['count'] for n in noise_names), 'reason': '名称像报告说明/无此项/判定语，不像检测项目；请人工确认是否排除'})

out = root/'outputs'/'019f4c2c-e544-7db3-a07e-ee8db394fef1'/'2026Q4_新增候选检测项_语义归并待确认.xlsx'
wb = Workbook()
ws = wb.active
ws.title = '建议归并组'
headers = ['归并组ID','建议统一检测项','组内候选项数','组内总出现次数','候选检测项','出现次数','示例报告号','示例货号','示例颜色','示例详情','归并理由','人工判断','确认后统一名称','备注']
ws.append(headers)
for g in sorted(groups, key=lambda x: -x['total']):
    for idx, name in enumerate(sorted(g['names'], key=lambda n: -item_by_name[n]['count'])):
        item = item_by_name[name]
        ex = (item.get('examples') or [{}])[0]
        ws.append([
            g['gid'] if idx == 0 else '',
            g['suggested'] if idx == 0 else '',
            len(g['names']) if idx == 0 else '',
            g['total'] if idx == 0 else '',
            name,
            item['count'],
            ex.get('report_no',''), ex.get('sku',''), ex.get('color',''), ex.get('detail',''),
            g['reason'] if idx == 0 else '',
            '', '', '',
        ])

ws2 = wb.create_sheet('候选项总表')
ws2.append(['候选检测项','出现次数','建议归并组ID','建议统一检测项','示例报告号','示例货号','示例颜色','示例详情','人工判断','确认后统一名称','备注'])
name_to_group = {}
for g in groups:
    for n in g['names']:
        name_to_group[n] = (g['gid'], g['suggested'])
for item in items:
    name = item['simple_item']
    gid, sug = name_to_group.get(name, ('', ''))
    ex = (item.get('examples') or [{}])[0]
    ws2.append([name, item['count'], gid, sug, ex.get('report_no',''), ex.get('sku',''), ex.get('color',''), ex.get('detail',''), '', '', ''])

for wsx in [ws, ws2]:
    fill = PatternFill('solid', fgColor='1E6F7A')
    thin = Side(style='thin', color='DCE5E8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in wsx[1]:
        c.fill = fill
        c.font = Font(color='FFFFFF', bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    for row in wsx.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = border
    widths = [12,28,12,14,30,10,22,14,16,70,38,14,24,24]
    for i in range(1, wsx.max_column+1):
        wsx.column_dimensions[get_column_letter(i)].width = widths[i-1] if i <= len(widths) else 18
    wsx.freeze_panes = 'A2'
    wsx.auto_filter.ref = wsx.dimensions

wb.save(out)
summary = {
    'output': str(out),
    'candidate_count': len(items),
    'group_count': len(groups),
    'grouped_candidate_count': len(name_to_group),
    'ungrouped_candidate_count': len(items) - len(name_to_group),
    'top_groups': [(g['gid'], g['suggested'], len(g['names']), g['total'], g['names'][:8]) for g in sorted(groups, key=lambda x: -x['total'])[:20]],
}
print(json.dumps(summary, ensure_ascii=False, indent=2))
