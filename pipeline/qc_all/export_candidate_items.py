import json, re
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

root = Path.cwd()
data = json.loads((root/'pipeline/qc_all/item_candidates.json').read_text(encoding='utf-8'))
out = root/'outputs'/'019f4c2c-e544-7db3-a07e-ee8db394fef1'/'2026Q4_新增候选检测项整理_候选清洗版.xlsx'

wb = Workbook()
ws = wb.active
ws.title = '新增候选检测项整理'
headers = ['候选检测项', '出现次数', '示例货号', '示例颜色', '源报告单号', 'PDF报告号', 'PDF链接', '示例详情', '建议处理']
ws.append(headers)
for item in data:
    examples = item.get('examples') or [{}]
    ex = examples[0]
    ws.append([
        item.get('simple_item',''), item.get('count',0), ex.get('sku',''), ex.get('color',''),
        ex.get('source_order_no',''), ex.get('report_no',''), ex.get('url',''), ex.get('detail',''), '待人工确认/归并'
    ])

header_fill = PatternFill('solid', fgColor='1E6F7A')
thin = Side(style='thin', color='DCE5E8')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in ws[1]:
    c.fill = header_fill
    c.font = Font(color='FFFFFF', bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = border
for i, width in enumerate([28,10,14,18,24,24,60,80,18], start=1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
wb.save(out)

bad_non_chinese = [x['simple_item'] for x in data if not re.search(r'[\u4e00-\u9fff]', x['simple_item'])]
bad_num_tail = [x['simple_item'] for x in data if re.search(r'\d', x['simple_item'])]
print(json.dumps({
    'output': str(out),
    'candidate_count': len(data),
    'non_chinese_candidates': len(bad_non_chinese),
    'candidates_with_digits': len(bad_num_tail),
    'top20': [(x['simple_item'], x['count']) for x in data[:20]],
    'digit_examples': bad_num_tail[:20],
}, ensure_ascii=False, indent=2))
