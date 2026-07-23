import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
WORK_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "V2"
OUTPUT_PATH = OUTPUT_DIR / "质检报告全量_2026Q4_纵向版_候选清洗版.xlsx"


TITLE_FILL = PatternFill("solid", fgColor="0F4C5C")
NOTE_FILL = PatternFill("solid", fgColor="E8F1F3")
HEADER_FILL = PatternFill("solid", fgColor="1E6F7A")
ALT_FILL = PatternFill("solid", fgColor="F3F8F9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
PASS_FILL = PatternFill("solid", fgColor="E7F4EA")
FAIL_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="DCE5E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(wb, title, note, headers, rows, freeze="C5", red_columns=None, green_columns=None):
    ws = wb.create_sheet(title)
    last_column = get_column_letter(len(headers))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = note
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = Font(color="315A64", size=10)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row_index, row in enumerate(rows, start=5):
        base_fill = ALT_FILL if row_index % 2 else WHITE_FILL
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row_index, col_index, value)
            cell.fill = base_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_index == 1:
                cell.number_format = "0"
            value_text = str(value or "")
            header = headers[col_index - 1]
            if red_columns and header in red_columns and "不符合" in value_text:
                cell.fill = FAIL_FILL
                cell.font = Font(color="C00000", bold=True)
            if green_columns and header in green_columns and value_text == "符合":
                cell.fill = PASS_FILL

    widths = {}
    for col_index, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_index in range(5, min(ws.max_row, 104) + 1):
            max_len = max(max_len, len(str(ws.cell(row_index, col_index).value or "")))
        widths[col_index] = min(max(max_len + 2, 10), 42)
    for col_index, width in widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[4].height = 36
    ws.auto_filter.ref = f"A4:{last_column}{max(ws.max_row, 4)}"
    return ws


def item_display(values):
    parts = []
    for value in values:
        display = value.get("display", "")
        report_no = value.get("report_no", "")
        if display and report_no:
            parts.append(f"{report_no}: {display}")
        elif display:
            parts.append(display)
    return "\n".join(dict.fromkeys(parts))


def main():
    data = json.loads((WORK_DIR / "trial_data.json").read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)

    view_headers = ["货号", "颜色", "检测项", "检测结果", "异常判定", "PDF报告号", "源报告单号", "PDF链接"]
    view_rows = []
    for row in data["summary_rows"]:
        for item in data["standard_items"]:
            values = row["items"].get(item, [])
            if not values:
                continue
            display = item_display(values)
            verdict = "不符合" if "不符合" in display else ""
            report_nos = "\n".join(dict.fromkeys(value.get("report_no", "") for value in values if value.get("report_no")))
            view_rows.append([row["sku"], row["color"], item, display, verdict, report_nos, row["source_order_no"], row["urls"]])
    write_sheet(
        wb,
        "检测结果查看",
        "一行代表一个货号+颜色+检测项；合格项显示实测值，不合格项标红。",
        view_headers,
        view_rows,
        red_columns={"检测结果", "异常判定"},
    )

    summary_headers = ["货号", "颜色", "源报告单号", "源表最后修改时间", "PDF报告号", "总体判定", "全部PDF链接"] + data["standard_items"]
    summary_rows = []
    for row in data["summary_rows"]:
        summary_rows.append([
            row["sku"], row["color"], row["source_order_no"], row["source_modified_time"],
            row["report_nos"], row["overall_result"], row["urls"],
            *[item_display(row["items"].get(item, [])) for item in data["standard_items"]],
        ])
    write_sheet(
        wb,
        "检测汇总",
        "一行代表一个货号+颜色；基础列之后为统一检测项横向结果。",
        summary_headers,
        summary_rows,
        red_columns=set(data["standard_items"]),
    )

    detail_headers = ["货号", "颜色", "源报告单号", "PDF报告号", "检测机构", "原始检测项", "简体检测项", "标准检测项", "子项", "实测值", "单位", "标准要求", "检测方法", "判定", "状态", "PDF链接"]
    detail_rows = [
        [row.get(key, "") for key in ["sku", "color", "source_order_no", "report_no", "institution", "raw_item", "simple_item", "standard_item", "subitem", "result", "unit", "requirement", "method", "verdict", "status", "url"]]
        for row in data["detail_rows"]
    ]
    write_sheet(
        wb,
        "检测明细",
        "完整数据底表，每个PDF、每个颜色、每个检测项一行。",
        detail_headers,
        detail_rows,
        freeze="E5",
        red_columns={"判定"},
        green_columns={"判定"},
    )

    mapping_headers = ["检测机构", "原始检测项", "简体检测项", "统一检测项", "归并说明", "状态", "首次出现报告号", "首次出现PDF链接"]
    mapping_rows = [
        [row.get(key, "") for key in ["institution", "raw_item", "simple_item", "standard_item", "merge_basis", "status", "first_report_no", "first_url"]]
        for row in data["mapping_rows"]
    ]
    write_sheet(wb, "项目映射", "展示检测项目原始名称、简体名称与统一检测项之间的映射。", mapping_headers, mapping_rows)

    source_headers = ["源表行号", "货号", "源表颜色原文", "源报告单号", "样品类型", "总体判定", "最后修改时间", "PDF链接", "主表采用", "采用颜色", "处理状态", "异常原因"]
    source_rows = [
        [row.get(key, "") for key in ["source_row", "sku", "color_raw", "order_no", "sample_type", "overall_result", "modified_time", "url", "selected", "selected_colors", "processing_status", "error_reason"]]
        for row in data["source_index"]
    ]
    write_sheet(wb, "源报告索引", "保留选中款号的全部源记录和拆分后的全部PDF链接。", source_headers, source_rows)

    error_headers = ["异常类型", "货号", "颜色", "源报告单号", "PDF报告号", "PDF链接", "原始检测项", "简体检测项", "建议统一检测项", "详情"]
    error_rows = [
        [row.get(key, "") for key in ["type", "sku", "color", "source_order_no", "report_no", "url", "raw_item", "simple_item", "suggested_item", "detail"]]
        for row in data["errors"]
    ]
    write_sheet(wb, "异常日志", "集中记录下载、识别、归并和结果冲突问题。", error_headers, error_rows)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(json.dumps({"output": str(OUTPUT_PATH), "sheets": wb.sheetnames, **data["stats"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
