import json
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

root = Path.cwd()
out = root / 'outputs' / '019f4c2c-e544-7db3-a07e-ee8db394fef1' / '质检报告全量_2026Q4_纵向版.xlsx'
wb = load_workbook(out, read_only=True, data_only=False)
expected_sheets = ['检测结果查看', '检测汇总', '检测明细', '项目映射', '源报告索引', '异常日志']
assert wb.sheetnames == expected_sheets, wb.sheetnames

def headers(sheet):
    ws = wb[sheet]
    return [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]

required = {
    '检测明细': ['状态'],
    '项目映射': ['首次出现报告号', '首次出现PDF链接'],
    '源报告索引': ['异常原因'],
    '异常日志': ['源报告单号', 'PDF报告号', '原始检测项', '简体检测项', '建议统一检测项'],
}
for sheet, cols in required.items():
    hs = headers(sheet)
    missing = [c for c in cols if c not in hs]
    assert not missing, (sheet, missing, hs)

main_sheets = ['检测结果查看', '检测汇总', '检测明细']
main_zero = {}
unspecified = {}
for sheet in main_sheets:
    hs = headers(sheet)
    sku_col = hs.index('货号') + 1
    color_col = hs.index('颜色') + 1
    zero = 0
    uns = 0
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if str(row[sku_col-1]).strip() == '0':
            zero += 1
        if row[color_col-1] == '未标明颜色':
            uns += 1
    main_zero[sheet] = zero
    unspecified[sheet] = uns
assert all(v == 0 for v in main_zero.values()), main_zero
assert any(v > 0 for v in unspecified.values()), unspecified

source_h = headers('源报告索引')
status_i = source_h.index('处理状态')
selected_i = source_h.index('主表采用')
sku_i = source_h.index('货号')
status_counts = Counter()
selected_zero = Counter()
generic_parse = 0
for row in wb['源报告索引'].iter_rows(min_row=5, values_only=True):
    status = row[status_i]
    status_counts[status] += 1
    if status == '解析异常':
        generic_parse += 1
    if str(row[sku_i]).strip() == '0':
        selected_zero[row[selected_i]] += 1
assert generic_parse == 0, generic_parse

mapping_h = headers('项目映射')
map_status_i = mapping_h.index('状态')
candidate_mapping = 0
for row in wb['项目映射'].iter_rows(min_row=5, values_only=True):
    if row[map_status_i] in ('新增候选', '待确认'):
        candidate_mapping += 1

err_h = headers('异常日志')
err_type_i = err_h.index('异常类型')
err_counts = Counter()
for row in wb['异常日志'].iter_rows(min_row=5, values_only=True):
    err_counts[row[err_type_i]] += 1

stats = json.loads((root/'pipeline/qc_all/trial_data.json').read_text(encoding='utf-8'))['stats']
result = {
    'output': str(out),
    'sheets': wb.sheetnames,
    'sheet_rows': {name: wb[name].max_row - 4 for name in wb.sheetnames},
    'sheet_cols': {name: wb[name].max_column for name in wb.sheetnames},
    'stats': stats,
    'main_zero_sku_rows': main_zero,
    'unspecified_color_rows': unspecified,
    'generic_parse_status_rows': generic_parse,
    'zero_sku_source_selected_counts': dict(selected_zero),
    'candidate_mapping_rows': candidate_mapping,
    'error_counts_top': dict(err_counts.most_common(12)),
    'status_counts': dict(status_counts),
}
print(json.dumps(result, ensure_ascii=False, indent=2))
