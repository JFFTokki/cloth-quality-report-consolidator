import json
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path.cwd()
OUT_DIR = ROOT / "outputs" / "019f4c2c-e544-7db3-a07e-ee8db394fef1"
ITEMS_PATH = ROOT / "tmp" / "qc_all" / "item_candidates.json"
LATEST_WB_PATH = max((ROOT / "V2").glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
OUT_PATH = OUT_DIR / "2026Q4_新增检测项待确认_按当前逻辑_附PDF链接.xlsx"


def pick(row, *names):
    for name in names:
        value = row.get(name)
        if value:
            return value
    return ""


items = json.loads(ITEMS_PATH.read_text(encoding="utf-8"))
item_by_name = {x["simple_item"]: x for x in items}

wb_src = load_workbook(LATEST_WB_PATH, read_only=True, data_only=True)
ws_err = wb_src["异常日志"]
headers = [c.value for c in next(ws_err.iter_rows(min_row=4, max_row=4))]
idx = {h: i for i, h in enumerate(headers)}

all_error_rows = []
examples_by_item = defaultdict(list)
counts_by_item = Counter()
for values in ws_err.iter_rows(min_row=5, values_only=True):
    if values[idx["异常类型"]] != "新增检测项待确认":
        continue
    name = str(values[idx["简体检测项"]] or values[idx["原始检测项"]] or "").strip()
    if not name:
        continue
    record = {
        "异常类型": values[idx["异常类型"]] or "",
        "货号": values[idx["货号"]] or "",
        "颜色": values[idx["颜色"]] or "",
        "源报告单号": values[idx["源报告单号"]] or "",
        "PDF报告号": values[idx["PDF报告号"]] or "",
        "PDF链接": values[idx["PDF链接"]] or "",
        "原始检测项": values[idx["原始检测项"]] or "",
        "简体检测项": name,
        "建议统一检测项": values[idx["建议统一检测项"]] or name,
        "详情": values[idx["详情"]] or "",
    }
    all_error_rows.append(record)
    counts_by_item[name] += 1
    if len(examples_by_item[name]) < 8:
        examples_by_item[name].append(record)

for name, data in item_by_name.items():
    if counts_by_item[name]:
        data["count"] = counts_by_item[name]
        data["examples"] = examples_by_item[name]

manual_groups = {
    "耐摩擦色牢度": ["干摩", "干摩擦", "湿摩", "湿摩擦", "摩擦色牢度", "耐摩擦色牢度"],
    "缝子纰裂程度": ["缝子纰裂", "接缝性能缝子纰裂", "现象判定接缝性能", "缝子纰裂程度", "子纰裂程", "接缝性能"],
    "沾色": ["沾色", "耐唾液色牢腈纶", "酸汗碱聚酯"],
    "变色": ["变色"],
    "防泼水性能": ["防泼水性能", "拒水性", "防水性能", "泼水"],
    "防油性能": ["防油性能", "拒油性"],
    "防污性能": ["防污性能", "耐沾污性"],
    "掉纤测试": ["掉纤测试", "掉纤维", "织物掉纤维"],
    "杂质含量": ["杂质", "杂质含量"],
    "陆禽毛含量": ["陆禽毛", "陆禽毛含量"],
    "绒丝+羽丝含量": ["绒丝羽丝", "绒丝羽丝含量", "绒丝+羽丝"],
    "重金属": ["重金属", "可萃取汞", "镉", "铅", "铬", "铜", "镍", "钴", "锑", "砷", "汞"],
    "附件锐利性": ["附件锐利尖端", "附件锐利边缘", "金属附件允许轻微掉漆", "尖端", "锐利"],
    "附件尺寸": ["附件尺寸"],
    "防风性能": ["防风性能"],
    "透湿性能": ["透湿性能"],
    "勾丝性能": ["勾丝性能"],
    "延伸值": ["横向延伸值", "直向延伸值", "延伸值"],
    "维护方法": ["维护方法", "维护方法无"],
    "羽绒气味/浊度": ["气味", "浊度"],
}

noise_keywords = [
    "没有此项",
    "岁以下儿童没有此项",
    "表中表示",
    "本项判定",
    "测定低限",
    "批号款号",
    "结论",
    "本机构",
    "识还应含有产品质量检验",
    "其他",
    "以下空白",
    "备注",
    "声明",
    "报告",
    "委托",
]

assigned = {}
groups = []


def add_group(std, names, reason):
    names = [n for n in dict.fromkeys(names) if n in item_by_name and n not in assigned]
    if len(names) < 2:
        return
    gid = f"G{len(groups) + 1:03d}"
    total = sum(item_by_name[n].get("count", 0) for n in names)
    for n in names:
        assigned[n] = gid
    groups.append({"gid": gid, "suggested": std, "names": names, "total": total, "reason": reason})


for std, seeds in manual_groups.items():
    matched = []
    for name in item_by_name:
        if any(name == seed or seed in name or name in seed for seed in seeds):
            matched.append(name)
    add_group(std, matched, "人工种子词匹配：名称包含相同核心检测语义")

amine_tokens = ["苯胺", "联苯胺", "甲苯胺", "萘胺", "氨基偶氮", "二氨基", "二甲基", "二甲氧基", "亚甲基二", "对氯苯胺", "三甲基苯胺"]
amines = [n for n in item_by_name if n not in assigned and any(t in n for t in amine_tokens)]
add_group("可分解致癌芳香胺染料-具体芳香胺物质", amines, "疑似同属禁用偶氮/芳香胺具体物质明细，建议作为同一大类下子项")


def norm(name):
    value = name
    for token in ["性能", "测试", "试验", "含量", "程度", "项目", "采购内控标准", "森马集团", "针织服装", "客户要求", "判定"]:
        value = value.replace(token, "")
    return value.replace("乾", "干").replace("溼", "湿").strip()


remaining = [n for n in item_by_name if n not in assigned and not any(k in n for k in noise_keywords)]
buckets = defaultdict(list)
for name in remaining:
    key = norm(name)
    if len(key) >= 2:
        buckets[key].append(name)
for names in buckets.values():
    add_group(min(names, key=len), names, "去除“性能/测试/含量/程度”等修饰词后名称一致")

remaining = [n for n in item_by_name if n not in assigned and not any(k in n for k in noise_keywords)]
used = set()
for i, a in enumerate(remaining):
    if a in used or a in assigned:
        continue
    cluster = [a]
    na = norm(a)
    for b in remaining[i + 1:]:
        if b in used or b in assigned:
            continue
        nb = norm(b)
        if min(len(na), len(nb)) < 3:
            continue
        ratio = SequenceMatcher(None, na, nb).ratio()
        contain = na in nb or nb in na
        if ratio >= 0.84 or (contain and min(len(na), len(nb)) >= 4):
            cluster.append(b)
    if len(cluster) >= 2:
        used.update(cluster)
        add_group(min(cluster, key=len), cluster, "文本相似度较高或核心词互相包含，需人工复核")

noise_names = [n for n in item_by_name if n not in assigned and any(k in n for k in noise_keywords)]
if noise_names:
    gid = f"G{len(groups) + 1:03d}"
    total = sum(item_by_name[n].get("count", 0) for n in noise_names)
    for name in noise_names:
        assigned[name] = gid
    groups.append({"gid": gid, "suggested": "疑似非检测项-建议排除", "names": noise_names, "total": total, "reason": "名称像报告说明/无此项/判定语，不像检测项目；请人工确认是否排除"})

wb = Workbook()
ws = wb.active
ws.title = "建议归并组"
ws.append(["归并组ID", "建议统一检测项", "组内候选项数", "组内总出现次数", "候选检测项", "出现次数", "PDF链接", "PDF报告号", "货号", "颜色", "示例详情", "归并理由", "人工判断", "确认后统一名称", "备注"])
for group in sorted(groups, key=lambda g: -g["total"]):
    names = sorted(group["names"], key=lambda n: (-item_by_name[n].get("count", 0), n))
    for pos, name in enumerate(names):
        data = item_by_name[name]
        ex = (examples_by_item.get(name) or data.get("examples") or [{}])[0]
        ws.append([
            group["gid"] if pos == 0 else "",
            group["suggested"] if pos == 0 else "",
            len(names) if pos == 0 else "",
            group["total"] if pos == 0 else "",
            name,
            data.get("count", 0),
            pick(ex, "PDF链接", "url"),
            pick(ex, "PDF报告号", "report_no"),
            pick(ex, "货号", "sku"),
            pick(ex, "颜色", "color"),
            pick(ex, "详情", "detail"),
            group["reason"] if pos == 0 else "",
            "",
            "",
            "",
        ])

ws2 = wb.create_sheet("未归组候选")
ws2.append(["候选检测项", "出现次数", "PDF链接", "PDF报告号", "货号", "颜色", "示例详情", "人工判断", "确认后统一名称", "备注"])
for name in sorted([n for n in item_by_name if n not in assigned], key=lambda n: (-item_by_name[n].get("count", 0), n)):
    data = item_by_name[name]
    ex = (examples_by_item.get(name) or data.get("examples") or [{}])[0]
    ws2.append([
        name,
        data.get("count", 0),
        pick(ex, "PDF链接", "url"),
        pick(ex, "PDF报告号", "report_no"),
        pick(ex, "货号", "sku"),
        pick(ex, "颜色", "color"),
        pick(ex, "详情", "detail"),
        "",
        "",
        "",
    ])

ws3 = wb.create_sheet("新增检测项明细")
ws3.append(["候选检测项", "货号", "颜色", "源报告单号", "PDF报告号", "PDF链接", "原始检测项", "详情"])
for row in all_error_rows:
    ws3.append([row["简体检测项"], row["货号"], row["颜色"], row["源报告单号"], row["PDF报告号"], row["PDF链接"], row["原始检测项"], row["详情"]])

ws4 = wb.create_sheet("统计")
ws4.append(["指标", "数量"])
ws4.append(["新增检测项待确认明细行数", len(all_error_rows)])
ws4.append(["候选检测项去重数", len(item_by_name)])
ws4.append(["建议归并组数", len(groups)])
ws4.append(["建议归并覆盖候选项数", len(assigned)])
ws4.append(["未归组候选项数", len(item_by_name) - len(assigned)])

fill = PatternFill("solid", fgColor="1E6F7A")
thin = Side(style="thin", color="DCE5E8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for sheet in wb.worksheets:
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

widths = {
    "建议归并组": [12, 28, 12, 14, 28, 10, 65, 22, 16, 18, 80, 38, 12, 24, 24],
    "未归组候选": [28, 10, 65, 22, 16, 18, 80, 12, 24, 24],
    "新增检测项明细": [28, 16, 18, 18, 22, 65, 28, 90],
    "统计": [32, 14],
}
for sheet in wb.worksheets:
    for i, width in enumerate(widths.get(sheet.title, []), 1):
        sheet.column_dimensions[get_column_letter(i)].width = width

OUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(json.dumps({
    "input_workbook": str(LATEST_WB_PATH),
    "output": str(OUT_PATH),
    "detail_rows": len(all_error_rows),
    "unique_candidates": len(item_by_name),
    "groups": len(groups),
    "grouped_candidates": len(assigned),
    "ungrouped_candidates": len(item_by_name) - len(assigned),
    "top_candidates": counts_by_item.most_common(20),
}, ensure_ascii=False, indent=2))
