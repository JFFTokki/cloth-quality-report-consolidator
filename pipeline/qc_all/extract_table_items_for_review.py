import json, re, sys
from collections import Counter, defaultdict
from pathlib import Path
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path('tmp/vendor').resolve()))
import pdfplumber

ROOT = Path.cwd()
WORK = ROOT / 'pipeline/qc_all'
OUT_DIR = ROOT / 'outputs' / '019f4c2c-e544-7db3-a07e-ee8db394fef1'
source = json.loads((WORK/'source_records.json').read_text(encoding='utf-8'))
manifest = json.loads((WORK/'download_manifest.json').read_text(encoding='utf-8'))
pdf_text = {d['url']: d for d in json.loads((WORK/'pdf_text.json').read_text(encoding='utf-8'))}

records_by_url = defaultdict(list)
for rec in source['selected_records']:
    for url in rec.get('urls', []):
        records_by_url[url].append(rec)

known_noise = {
    '序号', '检测项目', '检验项目', '检测结果', '单项判定', '单项评价', '项目描述', '标准要求', '实测值', '标准值', '单位', '检测方法',
    '备注', '结论', '检验结论', '检验结果', '样品信息', '产品信息', '判定依据', '检测类别', '委托检测', '委托送样',
}
noise_patterns = [
    r'^(序号|检测|检验)?项目$', r'^(标准|实测|检测|单项|评价|结果|备注).{0,8}$', r'.*报告.*', r'.*委托.*',
    r'.*样品.*', r'.*产品等级.*', r'.*安全类别.*', r'.*见下.*表.*', r'.*符合.*要求.*', r'.*本页.*', r'.*空白.*',
]

# Current known standard names from build_data.py RULES, used only for grouping suggestions.
build_text = (WORK/'build_data.py').read_text(encoding='utf-8')
known_standard_names = re.findall(r'\("([^"]+)", \[', build_text)

def clean_cell(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v).replace('\r','\n')).strip()

def compact(v):
    return re.sub(r'\s+', '', v or '')

def is_header_row(row):
    joined = ''.join(compact(c) for c in row)
    return ('检测项目' in joined or '检验项目' in joined) and ('单项' in joined or '检测结果' in joined or '检验结果' in joined or '结果' in joined)

def header_indices(row):
    idx = {'item': None, 'result': None, 'verdict': None}
    for i, cell in enumerate(row):
        c = compact(cell)
        if c in ('检测项目','检验项目') or ('检测项目' in c) or ('检验项目' in c):
            idx['item'] = i
        if c in ('检测结果','检验结果','实测结果') or ('检测结果' in c) or ('检验结果' in c):
            idx['result'] = i
        if ('单项判定' in c) or ('单项评价' in c) or c == '判定':
            idx['verdict'] = i
    return idx

def split_item_cell(text):
    text = clean_cell(text)
    if not text:
        return []
    # Front matter may list multiple commissioned test items in one cell.
    parts = re.split(r'[,，;；、]\s*|\n+', text)
    out = []
    for p in parts:
        p = p.strip(' ：:/')
        if p:
            out.append(p)
    return out or [text]

def normalize_item_name(name):
    name = clean_cell(name)
    name = re.sub(r'^\d{1,3}[.)、]?\s*', '', name)
    name = re.sub(r'\([^)]{0,40}\)|（[^）]{0,40}）|\[[^\]]{0,40}\]', '', name)
    name = re.split(r'GB/T|GB |FZ/T|QB/T|ISO|AATCC|ASTM|SN/T|EN ', name, maxsplit=1)[0]
    name = re.split(r'客户要求|采购内控标准|Q/|Semir|≤|≥|<|>|=|＝|N\.D\.|未检出|mg/kg|cm|mm|级', name, maxsplit=1, flags=re.I)[0]
    name = re.sub(r'^[\d\s,，.、\-]+', '', name)
    name = re.sub(r'[\d\s,，.、\-_%]+$', '', name)
    name = re.sub(r'\s+', '', name)
    return name

def valid_item(name):
    if not name or name in known_noise:
        return False
    if len(name) > 40:
        return False
    if not re.search(r'[\u4e00-\u9fff]', name):
        return False
    if re.search(r'\d', name):
        return False
    for pat in noise_patterns:
        if re.fullmatch(pat, name):
            return False
    if len(name) == 1 and name not in {'铅','镉','汞','砷','铬','镍'}:
        return False
    return True

def infer_verdict(result, verdict):
    text = f'{result} {verdict}'
    if re.search(r'不符合|不合格', text):
        return '不合格'
    if re.search(r'符合|合格|通过', text):
        return '合格'
    return verdict or ''

def get_report_no(url):
    d = pdf_text.get(url, {})
    return d.get('report_no') or Path(url.split('?')[0]).stem[:40]

rows = []
problems = Counter()
for m in manifest:
    url = m['url']
    if m.get('status') == 'failed' or not m.get('path'):
        continue
    path = Path(m['path'])
    if not path.exists():
        continue
    report_no = get_report_no(url)
    try:
        with pdfplumber.open(path) as pdf:
            for page_no, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables() or []
                for table_no, table in enumerate(tables, 1):
                    current = None
                    for rno, raw_row in enumerate(table, 1):
                        row = [clean_cell(c) for c in raw_row]
                        if is_header_row(row):
                            current = header_indices(row)
                            if current['item'] is None:
                                current = None
                            continue
                        if not current:
                            continue
                        item_i = current['item']
                        if item_i >= len(row):
                            continue
                        result = row[current['result']] if current.get('result') is not None and current['result'] < len(row) else ''
                        verdict = row[current['verdict']] if current.get('verdict') is not None and current['verdict'] < len(row) else ''
                        # If both selected result columns are empty, skip continuation rows.
                        if not result and not verdict:
                            continue
                        for raw_item in split_item_cell(row[item_i]):
                            item = normalize_item_name(raw_item)
                            if not valid_item(item):
                                problems['invalid_item_cell'] += 1
                                continue
                            linked_records = records_by_url.get(url) or [{}]
                            for rec in linked_records:
                                colors = rec.get('selected_colors') or ['未标明颜色']
                                for color in colors:
                                    rows.append({
                                        'item': item,
                                        'raw_item': raw_item,
                                        'result': result,
                                        'verdict': infer_verdict(result, verdict),
                                        'verdict_raw': verdict,
                                        'report_no': report_no,
                                        'url': url,
                                        'source_order_no': rec.get('order_no',''),
                                        'sku': rec.get('sku',''),
                                        'color': color,
                                        'page': page_no,
                                        'table': table_no,
                                    })
    except Exception as e:
        problems[f'pdf_error:{type(e).__name__}'] += 1

# De-duplicate exact rows caused by repeated tables/record links.
seen=set(); dedup=[]
for r in rows:
    key=(r['item'], r['result'], r['verdict_raw'], r['report_no'], r['url'], r['sku'], r['color'], r['page'])
    if key in seen: continue
    seen.add(key); dedup.append(r)
rows=dedup

# Suggest semantic groups from extracted item column only.
def group_key(name):
    s=name
    replacements = {
        '干摩':'耐摩擦色牢度', '干摩擦':'耐摩擦色牢度', '湿摩':'耐摩擦色牢度', '湿摩擦':'耐摩擦色牢度',
        '耐干摩擦色牢':'耐摩擦色牢度', '耐湿摩擦色牢':'耐摩擦色牢度', '摩擦色牢度':'耐摩擦色牢度',
        '缝子纰裂':'缝子纰裂程度', '接缝性能缝子纰裂':'缝子纰裂程度', '接缝性能':'缝子纰裂程度',
        '拒水性':'防泼水性能', '拒水性能':'防泼水性能', '防水性能':'防泼水性能', '防水性':'防泼水性能',
        '拒油性':'防油性能', '拒油性能':'防油性能', '耐沾污性':'防污性能',
        '陆禽毛':'陆禽毛含量', '杂质':'杂质含量', '绒丝羽丝':'绒丝+羽丝含量',
        '变色':'变色', '沾色':'沾色', '掉纤':'掉纤测试', '掉纤测试':'掉纤测试',
    }
    for k,v in replacements.items():
        if k in s:
            return v
    if any(t in s for t in ['苯胺','联苯胺','萘胺','氨基偶氮','二氨基','二甲基','二甲氧基','亚甲基二']):
        return '可分解致癌芳香胺染料-具体芳香胺物质'
    for std in known_standard_names:
        if s == std or s in std or std in s:
            return std
    return ''

item_stats = {}
for item, group in defaultdict(list).items():
    pass
by_item=defaultdict(list)
for r in rows:
    by_item[r['item']].append(r)

# Add fuzzy groups for still ungrouped exact-similar names.
suggested={name: group_key(name) for name in by_item}
ungrouped=[n for n,g in suggested.items() if not g]
used=set()
for i,a in enumerate(ungrouped):
    if a in used: continue
    cluster=[a]
    for b in ungrouped[i+1:]:
        if b in used: continue
        if min(len(a),len(b))>=3 and (a in b or b in a or SequenceMatcher(None,a,b).ratio()>=0.86):
            cluster.append(b)
    if len(cluster)>=2:
        label=min(cluster,key=len)
        for n in cluster:
            suggested[n]=label
            used.add(n)

groups=defaultdict(list)
for name, rs in by_item.items():
    groups[suggested.get(name) or '未归组'].append(name)

out = OUT_DIR/'2026Q4_按检测项目列抽取_同类项待判断.xlsx'
wb=Workbook()
ws=wb.active; ws.title='同类项待判断'
headers=['建议组ID','建议统一检测项','组内项目数','组内记录数','检测项目','项目记录数','示例检测结果','示例单项判定','示例PDF报告号','示例货号','示例颜色','PDF链接','人工判断','确认后统一名称','备注']
ws.append(headers)
gid=0
for label,names in sorted(groups.items(), key=lambda kv: -sum(len(by_item[n]) for n in kv[1])):
    if label=='未归组': continue
    gid+=1
    total=sum(len(by_item[n]) for n in names)
    group_id=f'G{gid:03d}'
    for idx,name in enumerate(sorted(names,key=lambda n:-len(by_item[n]))):
        rs=by_item[name]
        ex=rs[0]
        ws.append([group_id if idx==0 else '', label if idx==0 else '', len(names) if idx==0 else '', total if idx==0 else '', name, len(rs), ex['result'], ex['verdict_raw'], ex['report_no'], ex['sku'], ex['color'], ex['url'], '', '', ''])

ws2=wb.create_sheet('明细-检测项目列')
ws2.append(['检测项目','原始检测项目单元格','检测结果','单项判定','判定归一','PDF报告号','货号','颜色','源报告单号','PDF页码','PDF链接'])
for r in sorted(rows, key=lambda x:(x['item'], x['report_no'], x['sku'], x['color'])):
    ws2.append([r['item'], r['raw_item'], r['result'], r['verdict_raw'], r['verdict'], r['report_no'], r['sku'], r['color'], r['source_order_no'], r['page'], r['url']])

ws3=wb.create_sheet('未归组项目')
ws3.append(['检测项目','项目记录数','示例检测结果','示例单项判定','示例PDF报告号','示例货号','示例颜色','PDF链接','人工判断','确认后统一名称','备注'])
for name in sorted(groups.get('未归组',[]), key=lambda n:-len(by_item[n])):
    ex=by_item[name][0]
    ws3.append([name,len(by_item[name]),ex['result'],ex['verdict_raw'],ex['report_no'],ex['sku'],ex['color'],ex['url'],'','',''])

for wsx in wb.worksheets:
    fill=PatternFill('solid', fgColor='1E6F7A')
    thin=Side(style='thin', color='DCE5E8')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    for c in wsx[1]:
        c.fill=fill; c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=border
    for row in wsx.iter_rows(min_row=2):
        for c in row:
            c.alignment=Alignment(vertical='top',wrap_text=True); c.border=border
    widths=[12,30,12,12,28,12,35,16,22,14,16,70,14,24,24]
    for i in range(1, wsx.max_column+1):
        wsx.column_dimensions[get_column_letter(i)].width=widths[i-1] if i<=len(widths) else 18
    wsx.freeze_panes='A2'; wsx.auto_filter.ref=wsx.dimensions
wb.save(out)

summary={
    'output': str(out),
    'extracted_rows': len(rows),
    'unique_items': len(by_item),
    'suggested_groups': gid,
    'ungrouped_items': len(groups.get('未归组',[])),
    'problems': dict(problems),
    'top_items': Counter({k:len(v) for k,v in by_item.items()}).most_common(30),
}
print(json.dumps(summary, ensure_ascii=False, indent=2))
