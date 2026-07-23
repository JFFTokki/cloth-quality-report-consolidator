import json,re
from pathlib import Path
from openpyxl import load_workbook
out=Path('outputs/019f4c2c-e544-7db3-a07e-ee8db394fef1/2026Q4_按检测项目列抽取_同类项待判断.xlsx')
wb=load_workbook(out, read_only=True, data_only=True)
summary={'exists':out.exists(),'size':out.stat().st_size,'sheets':wb.sheetnames,'sheet_rows':{},'sheet_cols':{}}
for ws in wb.worksheets:
    summary['sheet_rows'][ws.title]=ws.max_row-1
    summary['sheet_cols'][ws.title]=ws.max_column
ws=wb['明细-检测项目列']
headers=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
item_i=headers.index('检测项目')
url_i=headers.index('PDF链接')
missing_url=0; digit_item=0; total=0
for row in ws.iter_rows(min_row=2, values_only=True):
    total+=1
    item=str(row[item_i] or '')
    url=str(row[url_i] or '')
    if not url: missing_url+=1
    if re.search(r'\d', item) or not re.search(r'[\u4e00-\u9fff]', item): digit_item+=1
summary.update({'detail_rows_checked':total,'missing_pdf_link_rows':missing_url,'digit_or_non_chinese_item_rows':digit_item})
print(json.dumps(summary,ensure_ascii=False,indent=2))
