import json
from collections import defaultdict, Counter
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

root = Path.cwd()
input_path = root/'outputs'/'019f4c2c-e544-7db3-a07e-ee8db394fef1'/'2026Q4_新增候选检测项_语义归并待确认.xlsx'
md_path = root/'docs'/'03_检测项目归并规则.md'
blank_out = root/'outputs'/'019f4c2c-e544-7db3-a07e-ee8db394fef1'/'2026Q4_新增候选检测项_待再次判断_附出处.xlsx'

wb = load_workbook(input_path, data_only=True)
ws = wb['建议归并组']
headers = [c.value for c in ws[1]]
idx = {h:i for i,h in enumerate(headers)}
current = {'gid':'', 'suggested':'', 'reason':''}
confirmed = []
excluded = []
blank = []
rename_values = {'同意','排除','不同意','拆分',''}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[idx['归并组ID']]: current['gid'] = str(row[idx['归并组ID']]).strip()
    if row[idx['建议统一检测项']]: current['suggested'] = str(row[idx['建议统一检测项']]).strip()
    if row[idx['归并理由']]: current['reason'] = str(row[idx['归并理由']]).strip()
    candidate = str(row[idx['候选检测项']] or '').strip()
    if not candidate:
        continue
    judgment = str(row[idx['人工判断']] or '').strip()
    confirm = str(row[idx['确认后统一名称']] or '').strip()
    if judgment == '排除':
        excluded.append((candidate, current.copy(), row))
    elif judgment == '同意' or confirm or (judgment and judgment not in rename_values):
        unified = confirm or (current['suggested'] if judgment == '同意' else judgment)
        confirmed.append({
            'gid': current['gid'], 'candidate': candidate, 'unified': unified,
            'suggested': current['suggested'], 'count': row[idx['出现次数']] or 0,
            'report_no': row[idx['示例报告号']] or '', 'sku': row[idx['示例货号']] or '',
            'color': row[idx['示例颜色']] or '', 'detail': row[idx['示例详情']] or '',
            'reason': current['reason'],
        })
    elif not judgment and not confirm:
        blank.append({'gid': current['gid'], 'suggested': current['suggested'], 'candidate': candidate,
                      'count': row[idx['出现次数']] or 0, 'report_no': row[idx['示例报告号']] or '',
                      'sku': row[idx['示例货号']] or '', 'color': row[idx['示例颜色']] or '',
                      'detail': row[idx['示例详情']] or '', 'reason': current['reason']})

by_unified = defaultdict(list)
for r in confirmed:
    by_unified[r['unified']].append(r)

# Replace previous generated section if rerun.
md = md_path.read_text(encoding='utf-8')
start = '<!-- BEGIN 人工确认新增候选归并规则 -->'
end = '<!-- END 人工确认新增候选归并规则 -->'
if start in md and end in md:
    before = md.split(start)[0].rstrip()
    after = md.split(end, 1)[1].lstrip()
    md = before + '\n\n' + after

lines = []
lines.append(start)
lines.append('')
lines.append('## 7. 人工确认新增候选归并规则')
lines.append('')
lines.append('来源：`2026Q4_新增候选检测项_语义归并待确认.xlsx` 中 `人工判断=同意` 或人工直接填写统一名称的记录。')
lines.append('')
lines.append('- 以下规则已由人工确认，可写入后续归并逻辑。')
lines.append('- `人工判断=排除` 的候选项不进入检测项目归并规则。')
lines.append('- 未填写人工判断的候选项已另行导出，继续等待人工判断。')
lines.append('')
lines.append('| 统一检测项 | 候选/别名 | 归并说明 | 首次/示例出处 |')
lines.append('|---|---|---|---|')
for unified in sorted(by_unified):
    rows = sorted(by_unified[unified], key=lambda x: (-int(x['count'] or 0), x['candidate']))
    aliases = '、'.join(dict.fromkeys(r['candidate'] for r in rows))
    reasons = '；'.join(dict.fromkeys(r['reason'] for r in rows if r['reason']))
    examples = []
    for r in rows[:3]:
        parts = []
        if r['report_no']: parts.append(f"PDF报告号：{r['report_no']}")
        if r['sku']: parts.append(f"货号：{r['sku']}")
        if r['color']: parts.append(f"颜色：{r['color']}")
        if r['detail']: parts.append(f"示例：{str(r['detail'])[:80]}")
        examples.append('，'.join(parts))
    lines.append(f"| {unified} | {aliases} | {reasons or '人工确认归并'} | {'<br>'.join(examples)} |")
if excluded:
    lines.append('')
    lines.append('### 人工确认排除项')
    lines.append('')
    lines.append('| 候选项 | 排除说明 |')
    lines.append('|---|---|')
    for candidate, current, row in excluded:
        lines.append(f"| {candidate} | 人工判断为排除，不作为检测项目归并。 |")
lines.append('')
lines.append(end)
md = md.rstrip() + '\n\n' + '\n'.join(lines) + '\n'
md_path.write_text(md, encoding='utf-8')

# Export blank rows with source.
out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = '待再次判断'
out_headers = ['归并组ID','建议统一检测项','候选检测项','出现次数','示例报告号','示例货号','示例颜色','示例详情','归并理由','人工判断','确认后统一名称','备注']
out_ws.append(out_headers)
for r in blank:
    out_ws.append([r['gid'], r['suggested'], r['candidate'], r['count'], r['report_no'], r['sku'], r['color'], r['detail'], r['reason'], '', '', ''])
fill = PatternFill('solid', fgColor='1E6F7A')
thin = Side(style='thin', color='DCE5E8')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in out_ws[1]:
    c.fill = fill; c.font = Font(color='FFFFFF', bold=True); c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = border
for row in out_ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True); c.border = border
widths = [12,28,30,10,24,14,16,80,40,14,24,24]
for i,w in enumerate(widths,1): out_ws.column_dimensions[get_column_letter(i)].width = w
out_ws.freeze_panes='A2'; out_ws.auto_filter.ref=out_ws.dimensions
out_wb.save(blank_out)

print(json.dumps({
    'confirmed_rows': len(confirmed),
    'confirmed_unified_items': len(by_unified),
    'excluded_rows': len(excluded),
    'blank_rows_exported': len(blank),
    'md_path': str(md_path),
    'blank_output': str(blank_out),
}, ensure_ascii=False, indent=2))
