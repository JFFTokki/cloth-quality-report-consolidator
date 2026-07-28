import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import extract_table_items_checkpoint as extractor
from qc_rules import prefer_specific_detail_rows, to_simplified
from report_header import normalize_institution_display_value
from source_relationship import source_relationship_id
from pipeline_versions import (
    HEADER_PARSER_VERSION,
    MAPPING_RULE_VERSION,
    OCR_CONFIG_VERSION,
    TABLE_PARSER_VERSION,
    TEXT_EXTRACTOR_VERSION,
)


ROOT = Path.cwd()
WORK = Path(os.environ.get("QC_WORK_DIR", ROOT / "pipeline" / "qc_all")).resolve()
OUT_DIR = ROOT / "outputs" / "test_V2"
OUTPUT_PATH = Path(os.environ.get(
    "QC_OUTPUT_PATH",
    OUT_DIR / "2026Q4_随机100款_解析增强版_20260715.xlsx",
)).resolve()

source = json.loads((WORK / "source_records.json").read_text(encoding="utf-8"))
SOURCE_WORKBOOK = source.get("source_workbook", "")
manifest = json.loads((WORK / "download_manifest.json").read_text(encoding="utf-8"))
pdf_docs = json.loads((WORK / "pdf_text.json").read_text(encoding="utf-8"))
manifest_by_url = {row["url"]: row for row in manifest}
doc_by_url = {row["url"]: row for row in pdf_docs}

forced_skus = [
    sku for sku in re.findall(r"(?<!\d)\d{12}(?!\d)", os.environ.get("QC_FORCE_INCLUDE_SKUS", ""))
    if sku in set(source.get("selected_skus", []))
]
only_skus = extractor.parse_requested_skus("QC_ONLY_SKUS", set(source.get("selected_skus", [])))
sample_limit_text = os.environ.get("QC_SAMPLE_LIMIT", "100").strip().lower()
sample_limit = None if sample_limit_text in {"", "0", "all"} else int(sample_limit_text)
sample_skus = list(dict.fromkeys(only_skus if os.environ.get("QC_ONLY_SKUS") is not None else (forced_skus + source["selected_skus"])))
if sample_limit is not None:
    sample_skus = sample_skus[:sample_limit]
sample_sku_set = set(sample_skus)
sample_records = [record for record in source["selected_records"] if record.get("sku") in sample_sku_set]
source_lookup = {}
for record in sample_records:
    for url in record.get("urls", []):
        source_lookup.setdefault((url, record.get("sku", "")), record)


def source_context(url, sku):
    return source_lookup.get((url, sku), {})


records_by_key = defaultdict(list)
for record in sample_records:
    for color in record.get("selected_colors") or ["未标明颜色"]:
        records_by_key[(record["sku"], color)].append(record)

sample_urls = []
for record in sample_records:
    for url in record.get("urls", []):
        if url not in sample_urls:
            sample_urls.append(url)


TITLE_FILL = PatternFill("solid", fgColor="0F4C5C")
NOTE_FILL = PatternFill("solid", fgColor="E8F1F3")
HEADER_FILL = PatternFill("solid", fgColor="1E6F7A")
ALT_FILL = PatternFill("solid", fgColor="F3F8F9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
PASS_FILL = PatternFill("solid", fgColor="E7F4EA")
FAIL_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="DCE5E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(wb, title, note, headers, rows, freeze="C5", red_columns=None, green_columns=None, nowrap_columns=None):
    ws = wb.create_sheet(title)
    last_column = get_column_letter(len(headers))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = note
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = Font(color="315A64", size=10)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row_index, row in enumerate(rows, start=5):
        base_fill = ALT_FILL if row_index % 2 else WHITE_FILL
        for col_index, value in enumerate(row, start=1):
            header = headers[col_index - 1]
            if header == "货号" and value not in (None, ""):
                value = str(value)
            cell = ws.cell(row_index, col_index, value)
            cell.fill = base_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=not (nowrap_columns and header in nowrap_columns))
            if header == "货号":
                cell.number_format = "@"
            elif header == "源表行号":
                cell.number_format = "0"
            value_text = str(value or "")
            if red_columns and header in red_columns and "不符合" in value_text:
                cell.fill = FAIL_FILL
                cell.font = Font(color="C00000", bold=True)
            if green_columns and header in green_columns and value_text == "符合":
                cell.fill = PASS_FILL
    for col_index, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_index in range(5, min(ws.max_row, 104) + 1):
            max_len = max(max_len, len(str(ws.cell(row_index, col_index).value or "")))
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 10), 42)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[4].height = 36
    ws.auto_filter.ref = f"A4:{last_column}{max(ws.max_row, 4)}"
    return ws


def item_display(values, show_report_count=True):
    grouped = defaultdict(list)
    for value in values:
        display = value.get("display", "")
        report_no = value.get("report_no", "")
        if display and report_no and report_no not in grouped[display]:
            grouped[display].append(report_no)
        elif display:
            grouped.setdefault(display, [])
    parts = []
    for display, report_nos in grouped.items():
        if len(report_nos) == 1:
            parts.append(f"{report_nos[0]}: {display}")
        elif report_nos and show_report_count:
            parts.append(f"{display}（{len(report_nos)}份报告）")
        elif report_nos:
            parts.append(display)
        else:
            parts.append(display)
    return "\n".join(parts)


def split_parent_child(item):
    if "-" not in item:
        return clean_part(item), ""
    # 化学物质名称中的连字符不是父子项分隔符。
    if re.match(r"^(?:N|O|P|\d[\d,]*)-", item, re.I):
        if item.count("-") == 1:
            return clean_part(item), ""
        parent, child = item.rsplit("-", 1)
        return clean_part(parent), clean_part(child)
    parent, child = item.split("-", 1)
    return clean_part(parent), clean_part(child)


def clean_part(value):
    value = str(value or "").strip()
    value = value.strip("-—－/／\\,，.、:：;；●•·▪■ ")
    return "" if value in {"/", "／", "\\", "-", "—", "－", ",", "，", "、", "●", "•", "·", "▪", "■"} else value


def normalize_item_text(value):
    """先清除PDF字体噪声并统一兼容字符，再进入检测项语义判断。"""
    value = unicodedata.normalize("NFKC", to_simplified(str(value or "")))
    value = "".join(
        char for char in value
        if unicodedata.category(char) not in {"Cc", "Cf", "Co", "Cs"}
    )
    value = re.sub(r"^[\s\-—–－:：;；,，.。·•●○◆◇■□▪▫※*+]+", "", value)
    value = re.sub(r"[\s·•●○◆◇■□▪▫※]+$", "", value)
    return clean_part(value)


TRAILING_UNIT_PATTERN = re.compile(
    r"^(.*?)[(\[]\s*(%|‰|cm|mm|m|N|cN|kN|Pa|kPa|MPa|mg/(?:kg|100g)|g/(?:m2|m²)|级)\s*[)\]]$",
    re.I,
)


def split_trailing_unit(value):
    item = normalize_item_text(value)
    match = TRAILING_UNIT_PATTERN.fullmatch(item)
    if not match:
        return item, ""
    return clean_part(match.group(1)), match.group(2)


def has_forbidden_unicode(value):
    return any(
        unicodedata.category(char) in {"Cc", "Cf", "Co", "Cs"}
        for char in str(value or "")
    )


def has_semantic_text(value):
    return bool(re.search(r"[\u4e00-\u9fffA-Za-z0-9]", str(value or "")))


def load_confirmed_item_map():
    """读取既有人工确认结果；新近明确规则在standardize_item中拥有更高优先级。"""
    path = ROOT / "outputs" / "manual_review" / "2026Q4_按检测项目列抽取_同类项待判断.xlsx"
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    output = {}
    for sheet_name in ("同类项待判断", "未归组项目"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "检测项目" not in headers or "确认后统一名称" not in headers:
            continue
        item_i = headers.index("检测项目")
        confirmed_i = headers.index("确认后统一名称")
        judgment_i = headers.index("人工判断") if "人工判断" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = clean_part(row[item_i])
            confirmed = clean_part(row[confirmed_i])
            judgment = clean_part(row[judgment_i]) if judgment_i is not None else ""
            if item and confirmed and judgment != "排除":
                output[item] = confirmed
    wb.close()
    return output


CONFIRMED_ITEM_MAP = load_confirmed_item_map()


def standardize_item(value):
    """归并机构命名差异，同时保留会改变检测含义的细项。"""
    item, _ = split_trailing_unit(value)
    item = re.sub(r"^[*'’′`]+", "", item).strip()
    item = clean_part(item)
    item = item.replace("色牢-度", "色牢度").replace("纰裂程-度", "纰裂程度")
    item = item.replace("分-解致", "分解致").replace("萃-取", "萃取")
    # PDF换行可能在完整检测核心词内部插入连字符；只修复明确的语义后缀，
    # 不影响“父项-材料/方向/条件”等真正的父子项分隔符。
    item = re.sub(
        r"(?<=[\u4e00-\u9fff])[-—－](?=(?:色牢度|摩擦(?:色牢度)?|强度|强力|性能|含量|程度|变化率|测定|分析|要求|外观)$)",
        "",
        item,
    )
    item = re.sub(r"^萃[-—－]", "可萃取重金属-", item)
    item = re.sub(r"^分解致癌芳香胺染料", "可分解致癌芳香胺染料", item)
    item = re.sub(r"^耐水色牢(?:度)?[-—－]?", "耐水色牢度-", item)
    item = re.sub(r"^耐皂洗色[-—－]?(?:牢度[-—－]?)?", "耐皂洗色牢度-", item)
    item = re.sub(r"^耐唾液色[-—－]?(?:牢度[-—－]?)?", "耐唾液色牢度-", item)
    item = re.sub(r"^耐酸汗渍[-—－]?(?:色牢度[-—－]?)?", "耐汗渍色牢度（酸汗）-", item)
    item = re.sub(r"^耐碱汗渍[-—－]?(?:色牢度[-—－]?)?", "耐汗渍色牢度（碱汗）-", item)
    item = re.sub(r"^耐湿摩擦[-—－]?(?:色牢度[-—－]?)?", "耐摩擦色牢度-湿摩-", item)
    item = re.sub(r"^面料耐干[-—－]?摩擦色牢度[-—－]?", "耐摩擦色牢度-干摩-", item)
    item = re.sub(r"^面料耐湿摩擦色牢度[-—－]?", "耐摩擦色牢度-湿摩-", item)
    item = re.sub(r"[-—－]+$", "", item)
    item = re.sub(r"^(?:ph|PH|Ph|pH)值", "pH值", item)
    item = re.sub(r"^pH值[（(]其他[）)]$", "pH值", item)
    if item.startswith("pH值"):
        suffix = clean_part(item[len("pH值"):])
        return "pH值" if suffix in {"", "其他"} else f"pH值-{suffix}"
    item = re.sub(r"^标识标签\*", "标识标签", item)
    if re.fullmatch(r"标识(?:标签)?\*?", item):
        return "标识标签"
    item = re.sub(r"^材质鉴别(?=$|[([])", "材质鉴定", item)
    item = re.sub(r"^\((绒丝\+羽丝)\)含量$", r"\1含量", item)
    if item == "可拆卸小附件":
        return "婴幼儿鞋小附件要求"
    if re.fullmatch(r"(?:面料起毛起球|面料起球(?:（梭织）)?|起球|起毛起球|抗起毛球)", item):
        return "抗起毛球"
    if item == "洗液沾色程度":
        return "洗液沾色程度"
    exact_aliases = {
        "耐水渍色牢度": "耐水色牢度", "耐水溃色牢度": "耐水色牢度",
        "酚黄变": "酚黄变色牢度",
        "异味测试": "异味", "附件抗拉强度": "附件抗拉强力",
        "干擦": "耐摩擦色牢度-干摩", "湿擦": "耐摩擦色牢度-湿摩",
        "拒水性（洗后）": "防泼水性能-洗后", "拒油性（洗后）": "防油性能-洗后",
        "拒油性能（拒油性）洗前": "防油性能-洗前",
        "羽绒蓬松度": "蓬松度", "掉纤": "掉纤维程度",
        "萃取重金属含量": "可萃取重金属", "种邻苯二甲酸酯的总含量": "邻苯二甲酸酯",
        "起毛起球性能": "抗起毛球", "标志、标签": "标识标签",
        "可触及的锐利尖端": "锐利尖端和锐利边缘", "可触及的锐利边缘和锐利尖端": "锐利尖端和锐利边缘",
        "锐利尖端和边缘": "锐利尖端和锐利边缘", "附件锐利尖端和锐利边缘": "锐利尖端和锐利边缘",
        "利尖端和边缘": "锐利尖端和锐利边缘", "锐利尖端": "锐利尖端和锐利边缘",
    }
    if item in exact_aliases:
        return exact_aliases[item]
    if item in CONFIRMED_ITEM_MAP:
        item = clean_part(CONFIRMED_ITEM_MAP[item])
    # 产品标准/机构名称是检测依据，不是检测项子项。
    item = re.sub(r"[-—－](?:森马集团[-—－])?(?:羽绒服装|针织服装|梭织服装|毛针织服装)$", "", item)
    item = re.sub(r"[-—－]见附表$", "", item)
    aliases = (
        (r"^(?:纤维成分及含量|纤维成分)$", "纤维含量"),
        (r"^甲醛$", "甲醛含量"),
        (r"^标识(?:标签)?\*?$", "标识标签"),
        (r"^(?:面料起球(?:（梭织）)?|起球|起毛起球)$", "抗起毛球"),
        (r"^(?:水洗后外观|水洗后外观质量|洗后外观质量)$", "洗后外观"),
        (r"^(?:接缝性能[（(]缝子纰裂程度[）)]|缝子纰裂)$", "缝子纰裂程度"),
        (r"^附件锐利性$", "锐利尖端和锐利边缘"),
        (r"^杂质$", "杂质含量"),
        (r"^陆禽毛$", "陆禽毛含量"),
        (r"^绒丝\+羽丝$", "绒丝+羽丝含量"),
    )
    for pattern, standard in aliases:
        if re.fullmatch(pattern, item):
            return standard
    item = re.sub(r"^可萃取重金属含量", "可萃取重金属", item)
    item = re.sub(r"^可萃取重金属[（(]mg[-/]*六价铬.*$", "可萃取重金属-六价铬", item, flags=re.I)
    item = re.sub(r"^耐湿摩擦色牢度", "耐摩擦色牢度-湿摩", item)
    item = re.sub(r"^耐干摩擦色牢度", "耐摩擦色牢度-干摩", item)
    item = re.sub(r"^耐干摩擦[-—－]?(?:色牢度[-—－]?)?", "耐摩擦色牢度-干摩-", item)
    item = re.sub(r"^耐湿摩擦[-—－]?(?:色牢度[-—－]?)?", "耐摩擦色牢度-湿摩-", item)
    item = re.sub(r"^(?:面料|里料|衬料)(?=耐(?:皂洗|水|汗渍|酸汗渍|碱汗渍|唾液|摩擦))", "", item)
    item = re.sub(r"^耐酸汗渍色牢度", "耐汗渍色牢度（酸汗）", item)
    item = re.sub(r"^耐碱汗渍色牢度", "耐汗渍色牢度（碱汗）", item)
    item = item.replace("耐汗渍色牢度〈酸汗）", "耐汗渍色牢度（酸汗）")
    if re.match(r"^可分解致[-—－]?癌芳香胺染料", item):
        item = re.sub(r"^可分解致[-—－]?癌芳香胺染料", "可分解致癌芳香胺染料", item)
        item = re.sub(r"[（(]mg/k[-—－]?g[）)]$", "", item, flags=re.I)
    if item.startswith(("水洗后外观质量-外观", "洗后外观质量-外观")):
        return "洗后外观"
    if item.startswith("纤维成分定性分析/定量分析"):
        return "纤维含量"
    if re.fullmatch(r"总和[（(]DINP\+DNOP\+DIDP[）)]", item, flags=re.I):
        return "邻苯二甲酸酯-总和(DINP+DNOP+DIDP)"
    if item == "干摩":
        return "耐摩擦色牢度-干摩"
    if item == "湿摩":
        return "耐摩擦色牢度-湿摩"
    item = re.sub(r"^(抗静电性能[-—－]电荷面密度\[洗(?:前|后)\]).*$", r"\1", item)
    item = re.sub(r"[-—－]?[（(]方法$", "", item)
    item = re.sub(r"[（(]N[-—－]", "-", item)
    item = item.replace("（", "-").replace("(", "-") if not re.search(r"[）)]", item) else item
    return clean_part(item)


AROMATIC_AMINES = {
    "氨基联苯", "联苯胺", "氯邻甲苯胺", "萘胺", "邻氨基偶氮甲苯", "邻氨基偶氮甲苯(o-AAT)",
    "硝基-邻甲苯胺", "对氯苯胺", "二氨基苯甲醚", "二氨基二苯甲烷", "二氨基二苯甲烷(MDA)",
    "二氯联苯胺", "二甲氧基联苯胺", "二甲基联苯胺", "二氨基二苯醚", "二氨基二苯醚(EDA)",
    "二氨基二苯硫醚", "邻甲苯胺", "二氨基甲苯", "二氨基甲苯(DAT)", "三甲基苯胺",
    "邻氨基苯甲醚", "氨基偶氮苯", "二甲基苯胺", "二氨基二苯硫醚", "氯邻甲基苯胺",
}


def _family_child(full_item, aliases):
    for alias in sorted(aliases, key=len, reverse=True):
        if full_item == alias:
            return ""
        if full_item.startswith(alias):
            suffix = full_item[len(alias):]
            if suffix.startswith(("-", "—", "－", "（", "(", "[", "［")):
                suffix = suffix.strip("-—－ ")
                if suffix.startswith(("（", "(")) and suffix.endswith(("）", ")")):
                    suffix = suffix[1:-1]
                suffix = suffix.replace("）（", "-").replace(")(", "-")
                suffix = suffix.strip("（()）[]［］ ")
                return clean_part(suffix)
    return None


def canonical_parent_child(raw_item, standardized_item):
    """统一到统计父项；材料、方向、洗前洗后、组分和化学物质作为子项保留。"""
    raw = clean_part(raw_item)
    full = clean_part(standardized_item)
    raw_no_code = re.sub(r"[（(][A-Za-z]{1,8}[）)]$", "", raw)
    if raw in AROMATIC_AMINES or raw_no_code in AROMATIC_AMINES or full in AROMATIC_AMINES:
        return "可分解致癌芳香胺染料", full if full in AROMATIC_AMINES else raw
    if re.match(r"^[■●•·▪]?N-亚硝基", raw, re.I) and raw != "N-亚硝基胺":
        return "N-亚硝基胺", raw.lstrip("■●•·▪")
    if re.match(r"^邻苯二甲酸(?:二|丁基|酯)", raw) and raw != "邻苯二甲酸酯":
        return "邻苯二甲酸酯", raw
    if re.fullmatch(r"总和[（(].*(?:DBP|BBP|DEHP|DINP|DNOP|DIDP).*[）)]", raw, re.I):
        return "邻苯二甲酸酯", raw
    metal_match = re.fullmatch(r"(铅|镉|汞|砷|铬|六价铬|镍|锑|钴|铜)(?:[（(].+[）)])?", raw)
    if metal_match:
        return "重金属", raw
    label_components = {
        "产品名称", "执行的产品标准", "产品号型或规格", "维护方法", "制造者的名称和地址",
        "使用和贮藏注意事项", "检验合格证明", "鞋号", "货号", "材质", "执行标准编号",
        "质量等", "产地", "三包规定", "生产日期", "颜色", "标识的形式及要求", "鞋-商标或企业名称",
    }
    if raw in label_components or raw.startswith("内包装（含吊牌）-"):
        return "标识标签", raw
    if raw.startswith("产品标识"):
        return "标识标签", clean_part(raw.removeprefix("产品标识"))
    condition_match = re.match(
        r"^(吸湿速干性|远红外发射率|远红外辐射温升|防紫外线性能|摩擦色牢度|干摩擦色牢度|湿摩擦色牢度|纤维含量)"
        r"(?:[（(\[]([^）)\]]+)[）)\]])?(?:[-—－](.+))?$",
        full,
    )
    if condition_match:
        parent, condition, suffix = condition_match.groups()
        child_parts = [part for part in (condition, suffix) if part]
        if parent in {"摩擦色牢度", "干摩擦色牢度", "湿摩擦色牢度"}:
            child_parts.insert(0, "干摩" if parent == "干摩擦色牢度" else "湿摩" if parent == "湿摩擦色牢度" else "")
            parent = "耐摩擦色牢度"
        return parent, clean_part("-".join(part for part in child_parts if part))
    if raw in {"附件锐利尖端", "附件锐利边缘"}:
        return "锐利尖端和锐利边缘", raw.removeprefix("附件")
    if raw in {"腰部", "背部", "肩带", "长袖袖口", "底边", "绳带"} or "绳带" in raw:
        return "绳带要求", raw
    if re.fullmatch(r"(?:试样的变色|沾色-?.*|醋纤(?:-.*)?|聚酯纤维(?:-.*)?|浅色沾色)", raw):
        return "色牢度项目待确认", raw
    if re.fullmatch(r"(?:聚酯纤维|锦纶|腈纶|羊毛|棉|醋纤)(?:-.*)?", raw):
        return "纤维含量", raw
    if raw in {"纬向", "横向", "直向", "面料", "里料", "面层", "底层", "金属附件", "要求", "条款"}:
        return "检测子项待确认", raw
    if re.search(
        r"^(?:分-测|符合GB|长至|大尺寸时|许打结|两端固定|带）|有长度|自由端|寸时|适的穿着|"
        r"除腰带|服装内|超出|脚踝|饰物|的绳|要求-头|伸-除|符合森马|长度-|端或-|应超-|"
        r"许-打结|头部和颈部不应|肩带应是|续且无自由端|的自由端)",
        raw,
    ):
        return "解析残片待确认", raw

    families = [
        ("pH值", ("pH值",)),
        ("纤维含量", ("纤维含量", "纤维成分及含量", "纤维成分定性分析/定量分析", "纤维成分定性分析", "纤维成分分析", "纤维成分")),
        ("甲醛含量", ("甲醛含量", "甲醛", "游离甲醛")),
        ("可分解致癌芳香胺染料", ("可分解致癌芳香胺染料", "可分解芳香胺染料测试", "禁用芳香胺")),
        ("可萃取重金属", ("可萃取重金属", "可萃取重金属含量")),
        ("重金属总量", ("重金属总量", "总铅和总镉含量")),
        ("邻苯二甲酸酯", ("邻苯二甲酸酯",)),
        ("耐皂洗色牢度", ("耐皂洗色牢度", "耐洗色牢度", "皂洗色牢度")),
        ("耐水色牢度", ("耐水色牢度",)),
        ("耐汗渍色牢度", ("耐汗渍色牢度（酸汗）", "耐汗渍色牢度（碱汗）", "耐汗渍色牢度", "耐酸汗渍色牢度", "耐碱汗渍色牢度")),
        ("耐摩擦色牢度", ("耐摩擦色牢度", "耐干摩擦色牢度", "耐湿摩擦色牢度")),
        ("耐唾液色牢度", ("耐唾液色牢度",)),
        ("耐光色牢度", ("耐光色牢度",)),
        ("耐光汗复合色牢度", ("耐光汗复合色牢度",)),
        ("酚黄变色牢度", ("酚黄变色牢度", "酚黄变")),
        ("拼接互染", ("拼接互染", "拼接互染色牢度")),
        ("羽绒成分测定", ("羽绒成分测定", "羽绒成分分析", "羽绒羽毛成分分析")),
        ("标识标签", ("标识标签", "标识")),
        ("材质鉴定", ("材质鉴定", "材质鉴别")),
        ("水洗尺寸变化率", ("水洗尺寸变化率",)),
        ("洗后外观", ("水洗后外观质量", "洗后外观质量", "洗后外观", "水洗后外观")),
        ("抗起毛球", ("抗起毛球", "起毛起球", "面料起球", "起球")),
        ("缝子纰裂程度", ("缝子纰裂程度", "缝子纰裂", "接缝性能(缝子纰裂程度)", "接缝性能（缝子纰裂程度）")),
        ("抗静电性能", ("抗静电性能",)),
        ("防污性能", ("防污性能", "耐沾污性")),
        ("防油性能", ("防油性能", "拒油性能", "拒油性")),
        ("防泼水性能", ("防泼水性能", "拒水性")),
        ("蓬松度", ("蓬松度",)),
        ("绒子含量", ("绒子含量",)),
        ("充绒量", ("充绒量",)),
        ("钻绒值", ("钻绒值",)),
        ("附件抗拉强力", ("附件抗拉强力", "附件抗拉强度")),
        ("绳带要求", ("绳带要求",)),
        ("锐利尖端和锐利边缘", ("锐利尖端和锐利边缘", "附件锐利性")),
        ("洗液沾色程度", ("洗液沾色程度", "洗液沾色")),
        ("勾丝性能", ("勾丝性能",)),
        ("耐贮存色牢度", ("耐贮存色牢度",)),
        ("重金属", ("重金属",)),
        ("烷基酚和烷基酚聚氧乙烯醚", ("烷基酚和烷基酚聚氧乙烯醚", "烷基酚")),
        ("防钻绒性", ("防钻绒性",)),
        ("透气性能", ("透气性能", "透气率")),
        ("延伸值", ("直向、横向延伸值", "直向延伸值", "横向延伸值")),
        ("含氯苯酚", ("含氯苯酚",)),
        ("N-亚硝基胺", ("N-亚硝基胺",)),
        ("拼接互染", ("深浅拼接沾色", "深浅拼接浅色沾色", "接互染程度")),
    ]
    for parent, aliases in families:
        child = _family_child(full, aliases)
        if child is not None:
            if parent == "耐汗渍色牢度":
                if full.startswith("耐酸汗渍色牢度"):
                    child = "酸汗" + (f"-{child}" if child else "")
                elif full.startswith("耐碱汗渍色牢度"):
                    child = "碱汗" + (f"-{child}" if child else "")
                elif full.startswith("耐汗渍色牢度（酸汗）"):
                    child = "酸汗" + (f"-{child}" if child and child != "酸汗" else "")
                elif full.startswith("耐汗渍色牢度（碱汗）"):
                    child = "碱汗" + (f"-{child}" if child and child != "碱汗" else "")
            return parent, clean_part(child)
    return full, ""


def semantic_role(row, raw_item):
    """判断名称在当前结果行中的角色；示例词只提供特征，不构成父项黑名单。"""
    if row.get("context_only"):
        return "context_fragment"
    raw = clean_part(raw_item)
    result = clean_part(row.get("result", ""))
    unit = clean_part(row.get("unit", ""))
    if not raw:
        return "fragment"
    if raw in {"续下页", "检验项目结束", "无无", "依据", "相关规定"} or re.search(r"(?:^注[:：]?检出限为$|检出限为$)", raw):
        return "fragment"
    if re.fullmatch(r"(?:含绒量(?:极限偏差)?|绒子含量|长毛片含量|异色毛绒(?:含量)?|绒丝\+羽丝(?:含量)?|绒丝|羽丝|杂质(?:含量)?|陆禽毛(?:含量)?)", raw):
        return "down_component"
    if re.search(
        r"^(?:分-测|符合(?:GB|森马)|长至|大尺寸时|许打结|两端固定|带）|有长度|自由端|寸时|"
        r"适的穿着|除腰带|服装内|超出|脚踝|饰物|的绳|要求-头|伸-除|长度-|端或-|应超-)",
        raw,
    ):
        return "fragment"
    if re.search(r"牢度|程度|含量|性能|强(?:度|力)|变化率|测定|分析|要求|外观", raw) and not re.search(r"[-—－](?:变色|沾色)", raw):
        return "parent_candidate"
    if re.search(r"(?:试样的)?变色|沾色", raw):
        return "result_dimension"
    # 方向词通常配合数值/等级结果出现；完整名称如“横向延伸值”仍可独立成为检测项。
    if len(raw) <= 6 and re.fullmatch(r"(?:经|纬|横|直|纵|斜|径|轴)(?:向|方向)", raw) and (result or unit):
        return "direction"
    # 短材料/部位/层次名称依赖同块父项；完整的“棉纤维含量”等不命中此规则。
    if len(raw) <= 12 and not re.search(r"含量|牢度|性能|强(?:度|力)|变化率|测定|分析|要求|外观|程度|值$", raw):
        if re.fullmatch(r"(?:[\u4e00-\u9fff]{1,8}(?:纤维|纶|毛|棉|纤|料|层|附件|革|毛皮|橡胶|塑料|帮面|鞋底|缝线|装饰件))(?:[、，,/+-][\u4e00-\u9fff]{1,8})*", raw):
            return "material_or_part"
    if raw in {"要求", "条款"}:
        return "structural_label"
    return "parent_candidate"


METHOD_PARENT_RULES = (
    (r"GB/T\s*3922", "耐汗渍色牢度"),
    (r"GB/T\s*5713", "耐水色牢度"),
    (r"GB/T\s*3920", "耐摩擦色牢度"),
    (r"GB/T\s*18886", "耐唾液色牢度"),
    (r"GB/T\s*11047", "勾丝性能"),
    (r"GB/T\s*21294", "缝子纰裂程度"),
    (r"GB/T\s*32008", "耐贮存色牢度"),
    (r"GB/T\s*17592", "可分解致癌芳香胺染料"),
    (r"GB/T\s*19942", "可分解致癌芳香胺染料"),
    (r"GB/T\s*19941\.2", "甲醛含量"),
)


def method_parent(method):
    compact = re.sub(r"\s+", "", str(method or "")).replace("—", "-").replace("–", "-")
    for pattern, parent in METHOD_PARENT_RULES:
        if re.search(pattern.replace(r"\s*", ""), compact, re.I):
            return parent
    return ""


def page_section_parents(url, page_number):
    """从同页编号章节恢复父项；机构换版时依靠章节结构而不是项目别名穷举。"""
    document = doc_by_url.get(url) or {}
    page = next((p for p in document.get("pages", []) if str(p.get("page")) == str(page_number)), {})
    output = []
    for line in str(page.get("text") or "").splitlines():
        line = re.sub(r"\s+", " ", line).strip()
        match = re.match(r"^\d+\s+(.{2,80}?)(?:\s*[（(](?:GB|FZ|ISO|Q/|SN|AATCC)|\s+GB/T)", line, re.I)
        if not match:
            continue
        heading = clean_part(re.sub(r"[（(].*$", "", match.group(1)))
        standardized = standardize_item(heading)
        parent, _ = canonical_parent_child(heading, standardized)
        if parent and "待确认" not in parent and parent not in output:
            output.append(parent)
    return output


def repair_item_from_page(row, item):
    """仅在页内原文提供完整证据时修复被列切分截掉的项名前缀。"""
    match = re.fullmatch(r"值\[([^\]]+)\]", clean_part(item))
    if not match:
        return item
    document = doc_by_url.get(row.get("url", "")) or {}
    page = next((p for p in document.get("pages", []) if str(p.get("page")) == str(row.get("page"))), {})
    material = re.escape(match.group(1))
    if re.search(rf"pH\s*值\s*\[{material}\]", str(page.get("text") or ""), re.I):
        return f"pH值-{match.group(1)}"
    return item


def compatible_sections(role, sections):
    if role == "result_dimension":
        return [item for item in sections if "色牢度" in item or item in {"拼接互染", "洗液沾色程度"}]
    if role == "direction":
        return [item for item in sections if re.search(r"接缝|滑移|纰裂|勾丝|尺寸|延伸|伸长|强力", item)]
    if role == "down_component":
        return [item for item in sections if "羽绒" in item or "绒" in item]
    return sections


def direct_role_parent(role, raw_item, method):
    raw = clean_part(raw_item)
    if role == "result_dimension":
        if re.search(r"耐(?:干|湿)?摩擦", raw):
            return "耐摩擦色牢度"
        if re.search(r"(?:酸|碱)汗", raw):
            return "耐汗渍色牢度"
        if re.search(r"耐水", raw):
            return "耐水色牢度"
        if re.search(r"耐(?:皂洗|洗)", raw):
            return "耐皂洗色牢度"
        if re.search(r"耐唾液", raw):
            return "耐唾液色牢度"
    compact_method = re.sub(r"\s+", "", str(method or ""))
    if role == "material_or_part" and "8629" in compact_method:
        return "洗后外观"
    if role == "down_component":
        return "羽绒成分测定"
    return ""


def active_parent_compatible(role, parent):
    if role == "result_dimension":
        return "色牢度" in parent or parent in {"拼接互染", "洗液沾色程度", "色牢度"}
    if role == "direction":
        return parent in {"缝子纰裂程度", "勾丝性能", "水洗尺寸变化率", "延伸值"}
    return role in {"material_or_part", "structural_label", "down_component"}


def resolve_parent_child(row, raw_item, standardized_item, state):
    """以结构、方法、语义和结果形态综合确定父子项，并返回可审计证据。"""
    role = semantic_role(row, standardized_item)
    base_parent, base_child = canonical_parent_child(raw_item, standardized_item)
    if role == "context_fragment":
        return "续行待归位", raw_item, "续行待归位", 0, "续行已保留，但未找到可安全附加的同表父项/子项", False
    if role == "fragment" or base_parent == "解析残片待确认":
        return "解析残片", raw_item, "解析残片", 100, "文本形态为说明句、断句或OCR残片", False
    if role == "parent_candidate" and "待确认" not in base_parent:
        state["active_parent"] = base_parent
        state["active_table"] = row.get("table")
        state["active_method"] = row.get("method", "")
        return base_parent, base_child, "自动归并", 100, "名称本身可独立表达检测对象", True
    if standardized_item == "婴幼儿鞋小附件要求":
        state["active_parent"] = base_parent
        state["active_table"] = row.get("table")
        state["active_method"] = row.get("method", "")
        return base_parent, base_child, "自动归并", 100, "最新明确规范化规则已确定统一父项", True
    if "待确认" not in base_parent and (base_child or base_parent != standardized_item):
        state["active_parent"] = base_parent
        state["active_table"] = row.get("table")
        state["active_method"] = row.get("method", "")
        return base_parent, base_child, "自动归并", 95, "规范化名称已明确包含父项与子项关系", True

    evidence = []
    score = 15
    parent = direct_role_parent(role, raw_item, row.get("method", ""))
    if parent:
        score += 30
        evidence.append(f"名称结构或检测方法直接指向{parent}")
    if not parent:
        parent = method_parent(row.get("method", ""))
    if parent:
        if not evidence:
            score += 30
            evidence.append(f"检测方法指向{parent}")

    sections = compatible_sections(role, page_section_parents(row.get("url", ""), row.get("page", "")))
    if role == "result_dimension" and re.search(r"(?:试样的)?变色", raw_item):
        state["section_index"] = state.get("section_index", -1) + 1
    if sections:
        index = min(max(state.get("section_index", 0), 0), len(sections) - 1)
        structural_parent = sections[index]
        if not parent or parent == structural_parent:
            parent = structural_parent
            score += 80
            evidence.append(f"同页编号检测块及其标准号指向{parent}")
    same_table = state.get("active_table") == row.get("table")
    method_compatible = not row.get("method") or not state.get("active_method") or row.get("method") == state.get("active_method")
    if not parent and state.get("active_parent") and same_table and method_compatible and active_parent_compatible(role, state["active_parent"]):
        parent = state["active_parent"]
        score += 50
        evidence.append(f"继承同一检测块父项{parent}")
    if row.get("result") or row.get("unit") or row.get("requirement"):
        score += 5
        evidence.append("结果、单位或标准要求形态匹配")

    if parent:
        state["active_parent"] = parent
        state["active_table"] = row.get("table")
        state["active_method"] = row.get("method", "")
        child = clean_part(standardized_item if role == "down_component" else raw_item)
        if role == "down_component":
            child = {
                "绒丝+羽丝含量": "绒丝+羽丝",
                "杂质含量": "杂质",
                "陆禽毛含量": "陆禽毛",
                "异色毛绒含量": "异色毛绒",
            }.get(child, child)
        if score < 80:
            if child and role not in {"structural_label"}:
                status = "自动归并（中置信度）"
            else:
                child = f"具体类型未识别-{child}" if child else "具体类型未识别"
                status = "自动归入父项（子项类型未完全识别）"
        else:
            status = "自动归并"
        return parent, child, status, min(score, 100), "；".join(evidence), True

    # 能识别为色牢度结果维度，但具体色牢度类型缺失时，进入领域父项而非逐条人工确认。
    if role == "result_dimension":
        child = clean_part(raw_item) or "具体类型未识别"
        return "色牢度", f"具体类型未识别-{child}", "自动归入领域父项", score, "名称语义可确定为色牢度结果维度", True
    if role in {"direction", "material_or_part", "structural_label"}:
        return "待确认归属", clean_part(raw_item), "待确认归属", score, "可确定为依赖上文的子项，但当前报告结构未恢复父项", False
    if base_parent in {"检测子项待确认", "色牢度项目待确认", "解析残片待确认"}:
        return "待确认归属", clean_part(base_child or raw_item), "待确认归属", score, "规范化结果仍为依赖上文的占位项，禁止自动新增并进入总览", False
    return base_parent, base_child, "自动新增", 80, "名称完整且检测对象明确，自动新增规范父项", True


TRADITIONAL_REVIEW_CHARS = set("檢測檢驗項單評價標準實測結果變錦綸纖維復堿鹼濕乾幹後觀絨強脹斷鄰")


def needs_traditional_review(value):
    return any(char in TRADITIONAL_REVIEW_CHARS for char in str(value or ""))


def report_no(url):
    return (doc_by_url.get(url) or {}).get("report_no") or Path(url.split("?")[0]).stem[:40]


detail_rows = []
mapping = {}
source_index = []
errors = []
stats_total = Counter()
stats_by_url = {}
context_states = {}
for index, url in enumerate(sample_urls, 1):
    manifest_row = manifest_by_url.get(url)
    if not manifest_row:
        stats_total["missing_manifest"] += 1
        continue
    rows, stats = extractor.process_pdf(manifest_row)
    stats_by_url[url] = dict(stats)
    stats_total.update(stats)
    url_status, _ = extractor.classify_processing_status(
        manifest_row,
        stats,
        doc_by_url.get(url, {}),
        len(rows),
    )
    for row in rows:
        if row.get("sku") not in sample_sku_set:
            continue
        simple_item = normalize_item_text(repair_item_from_page(row, clean_part(row.get("item", ""))))
        _, inferred_unit = split_trailing_unit(simple_item)
        full_standard_item = standardize_item(simple_item)
        state_key = (url, row.get("sku", ""), row.get("color", ""))
        state = context_states.setdefault(state_key, {"section_index": -1, "active_parent": "", "active_table": None, "active_method": "", "last_page": None})
        if state.get("last_page") != row.get("page"):
            state["section_index"] = -1
            if page_section_parents(url, row.get("page", "")):
                state["active_parent"] = ""
                state["active_table"] = None
                state["active_method"] = ""
            state["last_page"] = row.get("page")
        standard_item, standard_subitem, merge_status, confidence, merge_evidence, include_horizontal = resolve_parent_child(
            row, simple_item, full_standard_item, state
        )
        source_record = source_context(url, row.get("sku", ""))
        subitem = standard_subitem
        item_status = merge_status if url_status in {"已解析", "已解析（OCR）"} else url_status
        detail = {
            "sku": row.get("sku", ""),
            "subcategory": row.get("subcategory", "") or source_record.get("subcategory", ""),
            "color": row.get("color", ""),
            "source_order_no": row.get("source_order_no", ""),
            "source_sheet": row.get("source_sheet", "") or source_record.get("source_sheet", ""),
            "source_row": row.get("source_row", "") or source_record.get("source_row", ""),
            "source_cell": row.get("source_cell", "") or source_record.get("source_cell", ""),
            "sample_type": row.get("sample_type", ""),
            "report_no": row.get("report_no", ""),
            "institution": normalize_institution_display_value((doc_by_url.get(url) or {}).get("institution", "")),
            "report_issue_date": (doc_by_url.get(url) or {}).get("report_issue_date", ""),
            "report_issue_date_status": (doc_by_url.get(url) or {}).get("report_issue_date_status", ""),
            "report_issue_date_reason": (doc_by_url.get(url) or {}).get("report_issue_date_reason", ""),
            "cma_mark": (doc_by_url.get(url) or {}).get("cma_mark", ""),
            "cnas_mark": (doc_by_url.get(url) or {}).get("cnas_mark", ""),
            "cma_recognition_note": (doc_by_url.get(url) or {}).get("cma_evidence", ""),
            "cnas_recognition_note": (doc_by_url.get(url) or {}).get("cnas_evidence", ""),
            "report_product_code": "\n".join((doc_by_url.get(url) or {}).get("report_product_codes", [])),
            "plate_number": "\n".join((doc_by_url.get(url) or {}).get("plate_numbers", [])),
            "material_number": "\n".join((doc_by_url.get(url) or {}).get("material_numbers", [])),
            "raw_item": row.get("raw_item", ""),
            "raw_row": row.get("raw_row", ""),
            "simple_item": simple_item,
            "full_standard_item": full_standard_item,
            "standard_item": standard_item,
            "subitem": subitem,
            "result": row.get("result", ""),
            "result_detail": row.get("result_detail", ""),
            "unit": inferred_unit if inferred_unit and (
                not clean_part(row.get("unit", ""))
                or re.search(r"[\u4e00-\u9fff]", str(row.get("unit", "")))
            ) else row.get("unit", ""),
            "cas_number": row.get("cas_number", ""),
            "detection_limit": row.get("detection_limit", ""),
            "requirement": row.get("requirement", ""),
            "method": row.get("method", ""),
            "verdict": row.get("verdict", ""),
            "page": row.get("page", ""),
            "status": item_status,
            "merge_confidence": confidence,
            "merge_evidence": merge_evidence,
            "include_horizontal": include_horizontal,
            "url": url,
            "metadata_diagnostics": json.dumps({
                "reportIssueDateLabel": (doc_by_url.get(url) or {}).get("report_issue_date_label", ""),
                "reportIssueDateOriginal": (doc_by_url.get(url) or {}).get("report_issue_date_original", ""),
                "reportIssueDateReason": (doc_by_url.get(url) or {}).get("report_issue_date_reason", ""),
                "reportIssueDateStatus": (doc_by_url.get(url) or {}).get("report_issue_date_status", ""),
                "reportIssueDateCandidates": (doc_by_url.get(url) or {}).get("report_issue_date_candidates", []),
                "institutionEvidence": (doc_by_url.get(url) or {}).get("institution_evidence", ""),
                "headerParserVersion": (doc_by_url.get(url) or {}).get("header_parser_version", ""),
                "cmaEvidence": (doc_by_url.get(url) or {}).get("cma_evidence", ""),
                "cnasEvidence": (doc_by_url.get(url) or {}).get("cnas_evidence", ""),
            }, ensure_ascii=False, separators=(",", ":")),
        }
        detail["source_relationship_id"] = source_relationship_id(
            workbook=SOURCE_WORKBOOK,
            sheet=detail["source_sheet"],
            row=detail["source_row"],
            cell=detail["source_cell"],
            url=url,
            sku=detail["sku"],
            color=detail["color"],
            sample_type=detail["sample_type"],
        )
        detail_rows.append(detail)
    if index % 50 == 0 or index == len(sample_urls):
        print(f"progress pdf={index}/{len(sample_urls)} detail_rows={len(detail_rows)}", flush=True)

detail_rows_before_preference = len(detail_rows)
detail_rows = prefer_specific_detail_rows(detail_rows)
stats_total["summary_only_duplicate_rows_removed"] += detail_rows_before_preference - len(detail_rows)
stats_total["rows_before_business_dedup"] = stats_total.get("rows", detail_rows_before_preference)
stats_total["rows"] = len(detail_rows)
rows_by_url_after_preference = Counter(row["url"] for row in detail_rows)
for url, url_stats in stats_by_url.items():
    url_stats["rows_before_business_dedup"] = url_stats.get("rows", 0)
    url_stats["rows"] = rows_by_url_after_preference[url]
mapping = {}
for detail in detail_rows:
    mapping.setdefault(
        (detail["institution"], detail["raw_item"], detail["standard_item"], detail["subitem"]),
        {
            "institution": detail["institution"],
            "raw_item": detail["raw_item"],
            "simple_item": detail["simple_item"],
            "full_standard_item": detail["full_standard_item"],
            "standard_item": detail["standard_item"],
            "standard_subitem": detail["subitem"],
            "merge_basis": detail["merge_evidence"],
            "merge_confidence": detail["merge_confidence"],
            "include_horizontal": "是" if detail["include_horizontal"] else "否",
            "status": detail["status"],
            "first_report_no": detail["report_no"],
            "first_url": detail["url"],
        },
    )

for record in sample_records:
    selected_colors = record.get("selected_colors") or ["未标明颜色"]
    for url in record.get("urls", []):
        if url not in sample_urls:
            continue
        selected = "是"
        parsed_count = sum(1 for row in detail_rows if row["url"] == url and row["sku"] == record["sku"])
        status, reason = extractor.classify_processing_status(
            manifest_by_url.get(url, {}),
            stats_by_url.get(url, {}),
            doc_by_url.get(url, {}),
            parsed_count,
        )
        metadata = extractor.processing_metadata(status, stats_by_url.get(url, {}), doc_by_url.get(url, {}))
        for source_color in selected_colors:
            relationship_id = source_relationship_id(
                workbook=SOURCE_WORKBOOK,
                sheet=record.get("source_sheet", ""),
                row=record.get("source_row", ""),
                cell=record.get("source_cell", ""),
                url=url,
                sku=record.get("sku", ""),
                color=source_color,
                sample_type=record.get("sample_type", ""),
            )
            if status not in {"已解析", "已解析（OCR）"}:
                errors.append({
                    "type": status,
                    "reason_code": metadata["reason_code"],
                    "sku": record["sku"],
                    "subcategory": record.get("subcategory", ""),
                    "color": source_color,
                    "source_order_no": record.get("order_no", ""),
                    "source_sheet": record.get("source_sheet", ""),
                    "source_row": record.get("source_row", ""),
                    "source_cell": record.get("source_cell", ""),
                    "source_relationship_id": relationship_id,
                    "sample_type": record.get("sample_type", ""),
                    "report_no": report_no(url),
                    "url": url,
                    "raw_item": "",
                    "simple_item": "",
                    "suggested_item": "",
                    "page": "",
                    "parse_method": metadata["parse_method"] or ("下载阶段" if status == "下载失败" else ""),
                    "diagnostic_metrics": metadata["diagnostic_metrics"],
                    "detail": reason,
                })
            source_index.append({
                "source_sheet": record.get("source_sheet", ""),
                "source_row": record.get("source_row", ""),
                "source_cell": record.get("source_cell", ""),
                "source_relationship_id": relationship_id,
                "sku": record.get("sku", ""),
                "subcategory": record.get("subcategory", ""),
                "color_raw": record.get("color_raw", ""),
                "color": source_color,
                "order_no": record.get("order_no", ""),
                "sample_type": record.get("sample_type", ""),
                "overall_result": record.get("overall_result", ""),
                "modified_time": record.get("modified_time", ""),
                "url": url,
                "selected": selected,
                "selected_colors": source_color,
                "parse_method": metadata["parse_method"],
                "processing_status": status,
                "reason_code": metadata["reason_code"],
                "error_reason": reason,
                "needs_review": metadata["needs_review"],
            })

include_all_invalid_records = os.environ.get("QC_INCLUDE_ALL_INVALID_RECORDS") == "1"
invalid_records_for_scope = source.get("invalid_records", []) if include_all_invalid_records else [
    record for record in source.get("invalid_records", [])
    if record.get("is_selected") or record.get("sku") in sample_sku_set
]
for record in invalid_records_for_scope:
    colors = record.get("colors") or ["未标明颜色"]
    urls = record.get("urls") or [""]
    reason = record.get("invalid_reason") or "无有效货号"
    for url in urls:
        for source_color in colors:
            relationship_id = source_relationship_id(
                workbook=SOURCE_WORKBOOK,
                sheet=record.get("source_sheet", ""),
                row=record.get("source_row", ""),
                cell=record.get("source_cell", ""),
                url=url,
                sku=record.get("sku", ""),
                color=source_color,
                sample_type=record.get("sample_type", ""),
            )
            errors.append({
                "type": reason,
                "reason_code": "invalid_sku",
                "sku": record.get("sku", ""),
                "subcategory": record.get("subcategory", ""),
                "color": source_color,
                "source_order_no": record.get("order_no", ""),
                "source_sheet": record.get("source_sheet", ""),
                "source_row": record.get("source_row", ""),
                "source_cell": record.get("source_cell", ""),
                "source_relationship_id": relationship_id,
                "sample_type": record.get("sample_type", ""),
                "report_no": report_no(url) if url else "",
                "url": url,
                "raw_item": "",
                "simple_item": "",
                "suggested_item": "",
                "page": "",
                "parse_method": "",
                "diagnostic_metrics": f"source_row={record.get('source_row', '')}; colors={'，'.join(colors)}",
                "detail": reason,
            })
            source_index.append({
                "source_sheet": record.get("source_sheet", ""),
                "source_row": record.get("source_row", ""),
                "source_cell": record.get("source_cell", ""),
                "source_relationship_id": relationship_id,
                "sku": record.get("sku", ""),
                "subcategory": record.get("subcategory", ""),
                "color_raw": record.get("color_raw", ""),
                "color": source_color,
                "order_no": record.get("order_no", ""),
                "sample_type": record.get("sample_type", ""),
                "overall_result": record.get("overall_result", ""),
                "modified_time": record.get("modified_time", ""),
                "url": url,
                "selected": "否",
                "selected_colors": "",
                "parse_method": "",
                "processing_status": reason,
                "reason_code": "invalid_sku",
                "error_reason": reason,
                "needs_review": "是",
            })

summary_rows = []
for key in source["summary_keys"]:
    sku, color = key["sku"], key["color"]
    if sku not in sample_sku_set:
        continue
    rows_for_key = [row for row in detail_rows if row["sku"] == sku and row["color"] == color]
    records = records_by_key.get((sku, color), [])
    item_values = defaultdict(list)
    report_nos = []
    urls = []
    order_nos = []
    modified_times = []
    overall_results = []
    for record in records:
        order_nos.append(record.get("order_no", ""))
        modified_times.append(record.get("modified_time", ""))
        overall_results.append(record.get("overall_result", ""))
        for url in record.get("urls", []):
            urls.append(url)
            rn = report_no(url)
            if rn:
                report_nos.append(rn)
    for row in rows_for_key:
        if not row.get("include_horizontal"):
            continue
        display = row["result"] if row["verdict"] == "合格" else (f'{row["result"]}｜{row["verdict"]}' if row["result"] else row["verdict"])
        if not display and row["verdict"] == "合格":
            continue
        if row.get("subitem") and display:
            display = f'{row["subitem"]}：{display}'
        tagged = {"report_no": row["report_no"], "display": display, "subitem": row.get("subitem", "")}
        if tagged not in item_values[row["standard_item"]]:
            item_values[row["standard_item"]].append(tagged)
    summary_rows.append({
        "sku": sku,
        "color": color,
        "source_order_no": "\n".join(dict.fromkeys(order_nos)),
        "source_modified_time": "\n".join(dict.fromkeys(modified_times)),
        "report_nos": "\n".join(dict.fromkeys(report_nos)),
        "overall_result": "\n".join(dict.fromkeys(overall_results)),
        "urls": "\n".join(dict.fromkeys(urls)),
        "items": item_values,
    })

used_items = {row["standard_item"] for row in detail_rows if row.get("include_horizontal")}
standard_items = sorted(used_items)
mapping_rows = sorted(mapping.values(), key=lambda row: (row["standard_item"], row["institution"], row["raw_item"]))

if os.environ.get("QC_LEGACY_XLSX") != "1":
    output_path = Path(os.environ.get("QC_DATA_PATH", WORK / "report_data.json"))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    def report_metadata_for_url(url):
        doc = doc_by_url.get(url) or {}
        manifest_row = manifest_by_url.get(url) or {}
        if manifest_row.get("status") == "failed":
            return {
                "report_no": report_no(url),
                "report_issue_date": "",
                "report_issue_date_label": "",
                "report_issue_date_status": "待复核",
                "report_issue_date_original": "",
                "report_issue_date_reason": f"待复核｜PDF下载失败，无法识别报告签发日期：{manifest_row.get('error', '')}",
                "report_issue_date_candidates": [],
                "institution": "未识别",
                "institution_evidence": f"未识别｜PDF下载失败，无法读取报告出具机构：{manifest_row.get('error', '')}",
                "header_parser_version": "",
                "header_parse_status": "下载失败",
                "header_parse_error": manifest_row.get("error", ""),
                "cma_mark": "待复核",
                "cnas_mark": "待复核",
                "cma_evidence": f"待复核｜PDF下载失败，无法检查CMA标识：{manifest_row.get('error', '')}",
                "cnas_evidence": f"待复核｜PDF下载失败，无法检查CNAS标识：{manifest_row.get('error', '')}",
                "report_product_codes": [],
                "plate_numbers": [],
                "material_numbers": [],
                "path": manifest_row.get("path", ""),
            }
        metadata = {
            key: doc.get(key, "")
            for key in (
                "report_no", "report_issue_date", "report_issue_date_label",
                "report_issue_date_status", "report_issue_date_original", "report_issue_date_reason",
                "report_issue_date_candidates", "institution",
                "institution_evidence", "header_parser_version", "header_parse_status", "header_parse_error",
                "cma_mark", "cnas_mark", "cma_evidence", "cnas_evidence",
                "report_product_codes", "plate_numbers", "material_numbers", "path",
            )
        }
        metadata["institution"] = normalize_institution_display_value(metadata.get("institution"))
        if metadata["institution"] == "未识别" and not metadata.get("institution_evidence"):
            metadata["institution_evidence"] = "未识别｜缓存中没有可用机构文字；按业务规则不再转人工确认"
        if metadata.get("report_issue_date") and not metadata.get("report_issue_date_status"):
            metadata["report_issue_date_status"] = "已识别"
        if metadata.get("report_issue_date") and not metadata.get("report_issue_date_reason"):
            metadata["report_issue_date_reason"] = "已识别｜旧缓存已提取报告签发日期，未保留新版候选诊断"
        if not metadata.get("report_issue_date") and not metadata.get("report_issue_date_status"):
            metadata["report_issue_date_status"] = "待复核"
        if not metadata.get("report_issue_date") and not metadata.get("report_issue_date_reason"):
            metadata["report_issue_date_reason"] = "待复核｜旧缓存未保留新版签发日期候选诊断，需重跑extract_text.py或人工复核首页"
        return metadata
    payload = {
        "source_workbook": SOURCE_WORKBOOK,
        "source_warnings": source.get("source_warnings", []),
        "classification_default": "基础检测",
        "traditional_to_simplified_applied": True,
        "pipeline_versions": {
            "text_extractor": TEXT_EXTRACTOR_VERSION,
            "ocr_config": OCR_CONFIG_VERSION,
            "report_header": HEADER_PARSER_VERSION,
            "table_parser": TABLE_PARSER_VERSION,
            "mapping_rules": MAPPING_RULE_VERSION,
        },
        "sample_seed": source.get("sample_seed", ""),
        "sample_skus": sample_skus,
        "sample_records": len(sample_records),
        "sample_urls": sample_urls,
        "detail_rows": detail_rows,
        "summary_rows": summary_rows,
        "standard_items": standard_items,
        "mapping_rows": mapping_rows,
        "source_index": source_index,
        "errors": errors,
        "stats": dict(stats_total),
        "stats_by_url": stats_by_url,
        "report_metadata_by_url": {url: report_metadata_for_url(url) for url in sample_urls},
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({
        "output": str(output_path),
        "sample_skus": len(sample_skus),
        "sample_records": len(sample_records),
        "sample_pdfs": len(sample_urls),
        "detail_rows": len(detail_rows),
        "summary_rows": len(summary_rows),
        "standard_items": len(standard_items),
        "errors": len(errors),
        "stats": dict(stats_total),
        "error_types": dict(Counter(row["type"] for row in errors)),
    }, ensure_ascii=False, indent=2))
    sys.exit(0)

wb = Workbook()
wb.remove(wb.active)

view_headers = ["货号", "颜色", "检测父项", "检测细项", "检测项", "检测结果", "异常判定", "PDF报告号", "源报告单号", "PDF链接"]
view_rows = []
for row in summary_rows:
    for item in standard_items:
        values = row["items"].get(item, [])
        if not values:
            continue
        parent = item
        child = "\n".join(dict.fromkeys(value.get("subitem", "") for value in values if value.get("subitem")))
        display = item_display(values, show_report_count=False)
        verdict = "不符合" if "不符合" in display else ""
        report_nos = "\n".join(dict.fromkeys(value.get("report_no", "") for value in values if value.get("report_no")))
        view_rows.append([row["sku"], row["color"], parent, child, item, display, verdict, report_nos, row["source_order_no"], row["urls"]])
write_sheet(wb, "检测结果查看", "一行代表一个货号+颜色+检测项；本版使用新逻辑保留父级+细项。", view_headers, view_rows, red_columns={"检测结果", "异常判定"}, nowrap_columns={"PDF链接"})

summary_headers = ["货号", "颜色", "源报告单号", "源表最后修改时间", "PDF报告号", "总体判定", "全部PDF链接"] + standard_items
summary_rows_out = []
for row in summary_rows:
    summary_rows_out.append([
        row["sku"], row["color"], row["source_order_no"], row["source_modified_time"],
        row["report_nos"], row["overall_result"], row["urls"],
        *[item_display(row["items"].get(item, [])) for item in standard_items],
    ])
write_sheet(wb, "检测汇总", "一行代表一个货号+颜色；基础列之后为新逻辑检测项横向结果。", summary_headers, summary_rows_out, red_columns=set(standard_items), nowrap_columns={"全部PDF链接"})

detail_headers = ["货号", "颜色", "源报告单号", "PDF报告号", "PDF页码", "检测机构", "检测父项", "检测细项", "原始检测项", "简体检测项", "归并前标准项", "统一检测项", "标准子项", "实测值", "检测结果细则", "单位", "标准要求", "检测方法", "判定", "归并状态", "置信度", "归并证据", "进入横向主表", "PDF链接"]
detail_rows_out = []
for row in detail_rows:
    parent, child = row.get("standard_item", ""), row.get("subitem", "")
    detail_rows_out.append([
        row.get("sku", ""),
        row.get("color", ""),
        row.get("source_order_no", ""),
        row.get("report_no", ""),
        row.get("page", ""),
        row.get("institution", ""),
        parent,
        child,
        row.get("raw_item", ""),
        row.get("simple_item", ""),
        row.get("full_standard_item", ""),
        row.get("standard_item", ""),
        row.get("subitem", ""),
        row.get("result", ""),
        row.get("result_detail", ""),
        row.get("unit", ""),
        row.get("requirement", ""),
        row.get("method", ""),
        row.get("verdict", ""),
        row.get("status", ""),
        row.get("merge_confidence", ""),
        row.get("merge_evidence", ""),
        "是" if row.get("include_horizontal") else "否",
        row.get("url", ""),
    ])
write_sheet(wb, "检测明细", "完整数据底表；每个PDF、每个颜色、每个父级+细项检测项一行。", detail_headers, detail_rows_out, freeze="E5", red_columns={"判定"}, green_columns={"判定"}, nowrap_columns={"PDF链接"})

mapping_headers = ["检测机构", "原始检测项", "简体检测项", "归并前标准项", "统一检测项", "标准子项", "归并说明", "置信度", "进入横向主表", "状态", "首次出现报告号", "首次出现PDF链接"]
mapping_rows_out = [
    [row.get(key, "") for key in ["institution", "raw_item", "simple_item", "full_standard_item", "standard_item", "standard_subitem", "merge_basis", "merge_confidence", "include_horizontal", "status", "first_report_no", "first_url"]]
    for row in mapping_rows
]
write_sheet(wb, "项目映射", "展示检测项目原始名称、简体名称与新逻辑统一检测项之间的映射。", mapping_headers, mapping_rows_out, nowrap_columns={"首次出现PDF链接"})

source_headers = ["源表行号", "货号", "源表颜色原文", "源报告单号", "样品类型", "总体判定", "最后修改时间", "PDF链接", "主表采用", "采用颜色", "解析方式", "处理状态", "原因代码", "异常原因", "需人工复核"]
source_rows_out = [
    [row.get(key, "") for key in ["source_row", "sku", "color_raw", "order_no", "sample_type", "overall_result", "modified_time", "url", "selected", "selected_colors", "parse_method", "processing_status", "reason_code", "error_reason", "needs_review"]]
    for row in source_index
]
write_sheet(wb, "源报告索引", "保留100款样本的全部源记录和PDF链接。", source_headers, source_rows_out, nowrap_columns={"PDF链接"})

error_headers = ["异常类型", "原因代码", "货号", "颜色", "源报告单号", "PDF报告号", "PDF链接", "PDF页码", "解析方式", "原始检测项", "简体检测项", "建议统一检测项", "诊断指标", "详情"]
error_rows_out = [
    [row.get(key, "") for key in ["type", "reason_code", "sku", "color", "source_order_no", "report_no", "url", "page", "parse_method", "raw_item", "simple_item", "suggested_item", "diagnostic_metrics", "detail"]]
    for row in errors
]
write_sheet(wb, "异常日志", "集中记录100款样本中新逻辑仍未识别的PDF。", error_headers, error_rows_out, nowrap_columns={"PDF链接"})

traditional_review_rows = []
for row in detail_rows:
    parent, child = row.get("standard_item", ""), row.get("subitem", "")
    check_values = [parent, child, row.get("full_standard_item", ""), row.get("raw_item", ""), row.get("simple_item", "")]
    if any(needs_traditional_review(value) for value in check_values):
        traditional_review_rows.append([
            row.get("sku", ""),
            row.get("color", ""),
            parent,
            child,
            row.get("standard_item", ""),
            row.get("raw_item", ""),
            row.get("result", ""),
            row.get("verdict", ""),
            row.get("report_no", ""),
            row.get("url", ""),
        ])
traditional_headers = ["货号", "颜色", "检测父项", "检测细项", "标准检测项", "原始检测项", "检测结果", "判定", "PDF报告号", "PDF链接"]
write_sheet(wb, "繁体复核", "列出仍可能残留繁体字或繁简混写的检测项，供人工复核后补充转换规则。", traditional_headers, traditional_review_rows, nowrap_columns={"PDF链接"})

qa_headers = ["检查项", "结果"]
qa_rows = [
    ["样本款号数", len(sample_skus)],
    ["样本源记录数", len(sample_records)],
    ["样本PDF数", len(sample_urls)],
    ["检测明细行数", len(detail_rows)],
    ["检测汇总行数", len(summary_rows_out)],
    ["横向检测项列数", len(standard_items)],
    ["待确认归属明细行数", sum(1 for row in detail_rows if row.get("status") == "待确认归属")],
    ["自动归入父项但子项类型未完全识别", sum(1 for row in detail_rows if row.get("status") == "自动归入父项（子项类型未完全识别）")],
    ["自动归入领域父项", sum(1 for row in detail_rows if row.get("status") == "自动归入领域父项")],
    ["解析残片行数", sum(1 for row in detail_rows if row.get("status") == "解析残片")],
    ["依赖上文的子项误作独立父项", sum(1 for row in detail_rows if semantic_role(row, row.get("simple_item", "")) != "parent_candidate" and row.get("standard_item") == row.get("simple_item") and row.get("include_horizontal"))],
    ["尾部残留括号项目数", sum(1 for item in standard_items if item.endswith("(") or item.endswith("（"))],
    ["父项或细项为无意义符号行数", sum(1 for row in detail_rows if not has_semantic_text(row.get("standard_item")) or has_forbidden_unicode(row.get("standard_item")) or has_forbidden_unicode(row.get("subitem")))],
    ["前导标点或项目符号残留项目数", sum(1 for item in standard_items if has_forbidden_unicode(item) or str(item).lstrip().startswith(("，", ",", "、", "●", "•", "·", "▪", "■", "□", "◆", "◇")))],
    ["繁体复核行数", len(traditional_review_rows)],
    ["抽取统计", json.dumps(dict(stats_total), ensure_ascii=False)],
]
write_sheet(wb, "质量检查", "100款正式结构报告的质量检查。", qa_headers, qa_rows, freeze="A2")

OUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT_PATH)
print(json.dumps({
    "output": str(OUTPUT_PATH),
    "sheets": wb.sheetnames,
    "sample_skus": len(sample_skus),
    "sample_records": len(sample_records),
    "sample_pdfs": len(sample_urls),
    "detail_rows": len(detail_rows),
    "summary_rows": len(summary_rows_out),
    "standard_items": len(standard_items),
    "errors": len(errors),
    "stats": dict(stats_total),
}, ensure_ascii=False, indent=2))
