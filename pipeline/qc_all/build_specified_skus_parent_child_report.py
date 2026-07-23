import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import extract_table_items_checkpoint as extractor


ROOT = Path.cwd()
WORK = ROOT / "tmp" / "qc_all"
OUT_DIR = ROOT / "outputs" / "test_V2"
OUTPUT_PATH = OUT_DIR / "2026Q4_指定12款_新逻辑质检报告汇总_V4.xlsx"

source = json.loads((WORK / "source_records.json").read_text(encoding="utf-8"))
manifest = json.loads((WORK / "download_manifest.json").read_text(encoding="utf-8"))
pdf_docs = json.loads((WORK / "pdf_text.json").read_text(encoding="utf-8"))
manifest_by_url = {row["url"]: row for row in manifest}
doc_by_url = {row["url"]: row for row in pdf_docs}

specified = json.loads((WORK / "specified_skus_from_attachment.json").read_text(encoding="utf-8"))
sample_skus = specified["matched"]
missing_skus = specified.get("missing", [])
sample_sku_set = set(sample_skus)
sample_records = [record for record in source["selected_records"] if record.get("sku") in sample_sku_set]
records_by_key = defaultdict(list)
for record in sample_records:
    for color in record.get("selected_colors") or ["未标明颜色"]:
        records_by_key[(record["sku"], color)].append(record)

sample_urls = []
for record in sample_records:
    for url in record.get("urls", []):
        if url not in sample_urls:
            sample_urls.append(url)


TITLE_FILL = PatternFill("solid", fgColor="0F4C5C")
NOTE_FILL = PatternFill("solid", fgColor="E8F1F3")
HEADER_FILL = PatternFill("solid", fgColor="1E6F7A")
ALT_FILL = PatternFill("solid", fgColor="F3F8F9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
PASS_FILL = PatternFill("solid", fgColor="E7F4EA")
FAIL_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="DCE5E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(wb, title, note, headers, rows, freeze="C5", red_columns=None, green_columns=None):
    ws = wb.create_sheet(title)
    last_column = get_column_letter(len(headers))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = note
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = Font(color="315A64", size=10)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row_index, row in enumerate(rows, start=5):
        base_fill = ALT_FILL if row_index % 2 else WHITE_FILL
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row_index, col_index, value)
            cell.fill = base_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_index == 1:
                cell.number_format = "0"
            value_text = str(value or "")
            header = headers[col_index - 1]
            if red_columns and header in red_columns and "不符合" in value_text:
                cell.fill = FAIL_FILL
                cell.font = Font(color="C00000", bold=True)
            if green_columns and header in green_columns and value_text == "符合":
                cell.fill = PASS_FILL
    for col_index, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_index in range(5, min(ws.max_row, 104) + 1):
            max_len = max(max_len, len(str(ws.cell(row_index, col_index).value or "")))
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 10), 42)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[4].height = 36
    ws.auto_filter.ref = f"A4:{last_column}{max(ws.max_row, 4)}"
    return ws


def item_display(values):
    parts = []
    for value in values:
        display = value.get("display", "")
        report_no = value.get("report_no", "")
        if display and report_no:
            parts.append(f"{report_no}: {display}")
        elif display:
            parts.append(display)
    return "\n".join(dict.fromkeys(parts))


def split_parent_child(item):
    if "-" not in item:
        return clean_part(item), ""
    parent, child = item.split("-", 1)
    return clean_part(parent), clean_part(child)


def clean_part(value):
    value = str(value or "").strip()
    value = value.strip("-—－/／\\,，.、:：;；●•·▪■ ")
    return "" if value in {"/", "／", "\\", "-", "—", "－", ",", "，", "、", "●", "•", "·", "▪", "■"} else value


TRADITIONAL_REVIEW_CHARS = set("檢測檢驗項單評價標準實測結果變錦綸纖維復堿鹼濕乾幹後觀絨強脹斷鄰")


def needs_traditional_review(value):
    return any(char in TRADITIONAL_REVIEW_CHARS for char in str(value or ""))


def report_no(url):
    return (doc_by_url.get(url) or {}).get("report_no") or Path(url.split("?")[0]).stem[:40]


detail_rows = []
mapping = {}
source_index = []
errors = []
stats_total = Counter()
for index, url in enumerate(sample_urls, 1):
    manifest_row = manifest_by_url.get(url)
    if not manifest_row:
        stats_total["missing_manifest"] += 1
        continue
    rows, stats = extractor.process_pdf(manifest_row)
    stats_total.update(stats)
    for row in rows:
        if row.get("sku") not in sample_sku_set:
            continue
        detail = {
            "sku": row.get("sku", ""),
            "color": row.get("color", ""),
            "source_order_no": row.get("source_order_no", ""),
            "report_no": row.get("report_no", ""),
            "institution": (doc_by_url.get(url) or {}).get("institution", ""),
            "raw_item": row.get("raw_item", ""),
            "simple_item": row.get("item", ""),
            "standard_item": row.get("item", ""),
            "subitem": row.get("item_detail", ""),
            "result": row.get("result", ""),
            "result_detail": row.get("result_detail", ""),
            "unit": "",
            "requirement": "",
            "method": "",
            "verdict": row.get("verdict", ""),
            "status": "已解析",
            "url": url,
        }
        detail_rows.append(detail)
        mapping[(detail["institution"], detail["raw_item"], detail["standard_item"])] = {
            "institution": detail["institution"],
            "raw_item": detail["raw_item"],
            "simple_item": detail["simple_item"],
            "standard_item": detail["standard_item"],
            "merge_basis": "新逻辑按PDF表格父级+细项组合生成，细项不被大类吞并",
            "status": "已解析",
            "first_report_no": detail["report_no"],
            "first_url": url,
        }
    if index % 50 == 0 or index == len(sample_urls):
        print(f"progress pdf={index}/{len(sample_urls)} detail_rows={len(detail_rows)}", flush=True)

for record in sample_records:
    for url in record.get("urls", []):
        if url not in sample_urls:
            continue
        selected = "是"
        parsed_count = sum(1 for row in detail_rows if row["url"] == url and row["sku"] == record["sku"])
        status = "已解析" if parsed_count else "未识别检测项"
        reason = "" if parsed_count else "新逻辑表格抽取未识别到检测项目"
        if not parsed_count:
            errors.append({
                "type": "未识别检测项",
                "sku": record["sku"],
                "color": record.get("color_raw", "") or "未标明颜色",
                "source_order_no": record.get("order_no", ""),
                "report_no": report_no(url),
                "url": url,
                "raw_item": "",
                "simple_item": "",
                "suggested_item": "",
                "detail": reason,
            })
        source_index.append({
            "source_row": record.get("source_row", ""),
            "sku": record.get("sku", ""),
            "color_raw": record.get("color_raw", ""),
            "order_no": record.get("order_no", ""),
            "sample_type": record.get("sample_type", ""),
            "overall_result": record.get("overall_result", ""),
            "modified_time": record.get("modified_time", ""),
            "url": url,
            "selected": selected,
            "selected_colors": "，".join(record.get("selected_colors", [])),
            "processing_status": status,
            "error_reason": reason,
        })

summary_rows = []
for key in source["summary_keys"]:
    sku, color = key["sku"], key["color"]
    if sku not in sample_sku_set:
        continue
    rows_for_key = [row for row in detail_rows if row["sku"] == sku and row["color"] == color]
    records = records_by_key.get((sku, color), [])
    item_values = defaultdict(list)
    report_nos = []
    urls = []
    order_nos = []
    modified_times = []
    overall_results = []
    for record in records:
        order_nos.append(record.get("order_no", ""))
        modified_times.append(record.get("modified_time", ""))
        overall_results.append(record.get("overall_result", ""))
        for url in record.get("urls", []):
            urls.append(url)
            rn = report_no(url)
            if rn:
                report_nos.append(rn)
    for row in rows_for_key:
        display = row["result"] if row["verdict"] == "合格" else (f'{row["result"]}｜{row["verdict"]}' if row["result"] else row["verdict"])
        if not display and row["verdict"] == "合格":
            continue
        tagged = {"report_no": row["report_no"], "display": display}
        if tagged not in item_values[row["standard_item"]]:
            item_values[row["standard_item"]].append(tagged)
    summary_rows.append({
        "sku": sku,
        "color": color,
        "source_order_no": "\n".join(dict.fromkeys(order_nos)),
        "source_modified_time": "\n".join(dict.fromkeys(modified_times)),
        "report_nos": "\n".join(dict.fromkeys(report_nos)),
        "overall_result": "\n".join(dict.fromkeys(overall_results)),
        "urls": "\n".join(dict.fromkeys(urls)),
        "items": item_values,
    })

used_items = {row["standard_item"] for row in detail_rows}
standard_items = sorted(used_items)
mapping_rows = sorted(mapping.values(), key=lambda row: (row["standard_item"], row["institution"], row["raw_item"]))

wb = Workbook()
wb.remove(wb.active)

view_headers = ["货号", "颜色", "检测父项", "检测细项", "检测项", "检测结果", "异常判定", "PDF报告号", "源报告单号", "PDF链接"]
view_rows = []
for row in summary_rows:
    for item in standard_items:
        values = row["items"].get(item, [])
        if not values:
            continue
        parent, child = split_parent_child(item)
        display = item_display(values)
        verdict = "不符合" if "不符合" in display else ""
        report_nos = "\n".join(dict.fromkeys(value.get("report_no", "") for value in values if value.get("report_no")))
        view_rows.append([row["sku"], row["color"], parent, child, item, display, verdict, report_nos, row["source_order_no"], row["urls"]])
write_sheet(wb, "检测结果查看", "一行代表一个货号+颜色+检测项；本版使用新逻辑保留父级+细项。", view_headers, view_rows, red_columns={"检测结果", "异常判定"})

summary_headers = ["货号", "颜色", "源报告单号", "源表最后修改时间", "PDF报告号", "总体判定", "全部PDF链接"] + standard_items
summary_rows_out = []
for row in summary_rows:
    summary_rows_out.append([
        row["sku"], row["color"], row["source_order_no"], row["source_modified_time"],
        row["report_nos"], row["overall_result"], row["urls"],
        *[item_display(row["items"].get(item, [])) for item in standard_items],
    ])
write_sheet(wb, "检测汇总", "一行代表一个货号+颜色；基础列之后为新逻辑检测项横向结果。", summary_headers, summary_rows_out, red_columns=set(standard_items))

detail_headers = ["货号", "颜色", "源报告单号", "PDF报告号", "检测机构", "检测父项", "检测细项", "原始检测项", "简体检测项", "标准检测项", "子项", "实测值", "检测结果细则", "单位", "标准要求", "检测方法", "判定", "状态", "PDF链接"]
detail_rows_out = []
for row in detail_rows:
    parent, child = split_parent_child(row.get("standard_item", ""))
    detail_rows_out.append([
        row.get("sku", ""),
        row.get("color", ""),
        row.get("source_order_no", ""),
        row.get("report_no", ""),
        row.get("institution", ""),
        parent,
        child,
        row.get("raw_item", ""),
        row.get("simple_item", ""),
        row.get("standard_item", ""),
        row.get("subitem", ""),
        row.get("result", ""),
        row.get("result_detail", ""),
        row.get("unit", ""),
        row.get("requirement", ""),
        row.get("method", ""),
        row.get("verdict", ""),
        row.get("status", ""),
        row.get("url", ""),
    ])
write_sheet(wb, "检测明细", "完整数据底表；每个PDF、每个颜色、每个父级+细项检测项一行。", detail_headers, detail_rows_out, freeze="E5", red_columns={"判定"}, green_columns={"判定"})

mapping_headers = ["检测机构", "原始检测项", "简体检测项", "统一检测项", "归并说明", "状态", "首次出现报告号", "首次出现PDF链接"]
mapping_rows_out = [
    [row.get(key, "") for key in ["institution", "raw_item", "simple_item", "standard_item", "merge_basis", "status", "first_report_no", "first_url"]]
    for row in mapping_rows
]
write_sheet(wb, "项目映射", "展示检测项目原始名称、简体名称与新逻辑统一检测项之间的映射。", mapping_headers, mapping_rows_out)

source_headers = ["源表行号", "货号", "源表颜色原文", "源报告单号", "样品类型", "总体判定", "最后修改时间", "PDF链接", "主表采用", "采用颜色", "处理状态", "异常原因"]
source_rows_out = [
    [row.get(key, "") for key in ["source_row", "sku", "color_raw", "order_no", "sample_type", "overall_result", "modified_time", "url", "selected", "selected_colors", "processing_status", "error_reason"]]
    for row in source_index
]
write_sheet(wb, "源报告索引", "保留100款样本的全部源记录和PDF链接。", source_headers, source_rows_out)

error_headers = ["异常类型", "货号", "颜色", "源报告单号", "PDF报告号", "PDF链接", "原始检测项", "简体检测项", "建议统一检测项", "详情"]
error_rows_out = [
    [row.get(key, "") for key in ["type", "sku", "color", "source_order_no", "report_no", "url", "raw_item", "simple_item", "suggested_item", "detail"]]
    for row in errors
]
write_sheet(wb, "异常日志", "集中记录100款样本中新逻辑仍未识别的PDF。", error_headers, error_rows_out)

traditional_review_rows = []
for row in detail_rows:
    parent, child = split_parent_child(row.get("standard_item", ""))
    check_values = [parent, child, row.get("standard_item", ""), row.get("raw_item", ""), row.get("simple_item", "")]
    if any(needs_traditional_review(value) for value in check_values):
        traditional_review_rows.append([
            row.get("sku", ""),
            row.get("color", ""),
            parent,
            child,
            row.get("standard_item", ""),
            row.get("raw_item", ""),
            row.get("result", ""),
            row.get("verdict", ""),
            row.get("report_no", ""),
            row.get("url", ""),
        ])
traditional_headers = ["货号", "颜色", "检测父项", "检测细项", "标准检测项", "原始检测项", "检测结果", "判定", "PDF报告号", "PDF链接"]
write_sheet(wb, "繁体复核", "列出仍可能残留繁体字或繁简混写的检测项，供人工复核后补充转换规则。", traditional_headers, traditional_review_rows)

qa_headers = ["检查项", "结果"]
qa_rows = [
    ["指定款号数", len(sample_skus)],
    ["样本源记录数", len(sample_records)],
    ["样本PDF数", len(sample_urls)],
    ["检测明细行数", len(detail_rows)],
    ["检测汇总行数", len(summary_rows_out)],
    ["横向检测项列数", len(standard_items)],
    ["短词单独成项行数", sum(1 for row in detail_rows if row["standard_item"] in {"沾色", "变色", "干摩", "湿摩"})],
    ["尾部残留括号项目数", sum(1 for item in standard_items if item.endswith("(") or item.endswith("（"))],
    ["父项或细项为无意义符号行数", sum(1 for row in detail_rows if split_parent_child(row["standard_item"])[0] in {"", "/", "-", "—", "－"} or split_parent_child(row["standard_item"])[1] in {"/", "-", "—", "－"})],
    ["前导标点或项目符号残留项目数", sum(1 for item in standard_items if str(item).lstrip().startswith(("，", ",", "、", "●", "•", "·", "▪", "■")))],
    ["繁体复核行数", len(traditional_review_rows)],
    ["抽取统计", json.dumps(dict(stats_total), ensure_ascii=False)],
]
write_sheet(wb, "质量检查", "100款正式结构报告的质量检查。", qa_headers, qa_rows, freeze="A2")

OUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT_PATH)
print(json.dumps({
    "output": str(OUTPUT_PATH),
    "sheets": wb.sheetnames,
    "matched_skus": len(sample_skus),
    "missing_skus": len(missing_skus),
    "sample_records": len(sample_records),
    "sample_pdfs": len(sample_urls),
    "detail_rows": len(detail_rows),
    "summary_rows": len(summary_rows_out),
    "standard_items": len(standard_items),
    "errors": len(errors),
    "stats": dict(stats_total),
}, ensure_ascii=False, indent=2))
