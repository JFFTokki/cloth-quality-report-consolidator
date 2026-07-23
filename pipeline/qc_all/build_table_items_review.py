import json
import re
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path.cwd()
WORK = ROOT / 'pipeline/qc_all'
OUT_DIR = ROOT / 'outputs' / '019f4c2c-e544-7db3-a07e-ee8db394fef1'
ROWS_PATH = WORK / 'table_items_rows.jsonl'
STATE_PATH = WORK / 'table_items_state.json'
build_text = (WORK / 'build_data.py').read_text(encoding='utf-8')
known_standard_names = re.findall(r'\("([^"]+)", \[', build_text)

def load_rows():
    rows = []
    if not ROWS_PATH.exists():
        return rows
    with ROWS_PATH.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    seen = set()
    dedup = []
    for r in rows:
        key = (r.get('item'), r.get('result'), r.get('verdict_raw'), r.get('report_no'), r.get('url'), r.get('sku'), r.get('color'), r.get('page'))
        if key in seen:
            continue
        seen.add(key)
        dedup.append(r)
    return dedup

def group_key(name):
    if '-' in name:
        # Full parent-child items must not be collapsed to the broad parent item.
        return name
    replacements = {
        '干摩': '耐摩擦色牢度', '干摩擦': '耐摩擦色牢度', '湿摩': '耐摩擦色牢度', '湿摩擦': '耐摩擦色牢度',
        '耐干摩擦色牢': '耐摩擦色牢度', '耐湿摩擦色牢': '耐摩擦色牢度', '摩擦色牢度': '耐摩擦色牢度',
        '缝子纰裂': '缝子纰裂程度', '接缝性能缝子纰裂': '缝子纰裂程度', '接缝性能': '缝子纰裂程度',
        '拒水性': '防泼水性能', '拒水性能': '防泼水性能', '防水性能': '防泼水性能', '防水性': '防泼水性能',
        '拒油性': '防油性能', '拒油性能': '防油性能', '耐沾污性': '防污性能',
        '陆禽毛': '陆禽毛含量', '杂质': '杂质含量', '绒丝羽丝': '绒丝+羽丝含量',
        '掉纤': '掉纤测试', '掉纤测试': '掉纤测试',
    }
    for k, v in replacements.items():
        if k in name:
            return v
    if any(t in name for t in ['苯胺', '联苯胺', '萘胺', '氨基偶氮', '二氨基', '二甲基', '二甲氧基', '亚甲基二']):
        return '可分解致癌芳香胺染料-具体芳香胺物质'
    for std in known_standard_names:
        if name == std or name in std or std in name:
            return std
    return ''

def extract_item_detail(raw_item):
    raw = raw_item or ''
    details = []
    for match in re.finditer(r'[（(［\[]([^）)\]］]+)[）)］\]]', raw):
        value = re.sub(r'\s+', '', match.group(1))
        if not value:
            continue
        # 单位不作为检测项目细则；面料、部位、颜色、样品类型等保留。
        if re.fullmatch(r'[%/]|g|kg|mg/kg|cm|mm|根/m²|根/m2|级|个|N', value, flags=re.I):
            continue
        if value in {'其他'}:
            continue
        details.append(value)
    return '；'.join(dict.fromkeys(details))

def extract_result_detail(result):
    text = result or ''
    # 检测结果本身常带“变色/沾色/面料/部位 + 数值”，整体保留给人工判断。
    return re.sub(r'\s+', '\n', text).strip()

def style_sheet(ws):
    fill = PatternFill('solid', fgColor='1E6F7A')
    thin = Side(style='thin', color='DCE5E8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = fill
        c.font = Font(color='FFFFFF', bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = border
    widths = [12, 30, 12, 12, 28, 22, 12, 35, 35, 16, 22, 14, 16, 70, 14, 24, 24]
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i <= len(widths) else 18
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def main():
    rows = load_rows()
    by_item = defaultdict(list)
    for r in rows:
        by_item[r['item']].append(r)
    suggested = {name: group_key(name) for name in by_item}
    ungrouped = [n for n, g in suggested.items() if not g]
    used = set()
    for i, a in enumerate(ungrouped):
        if a in used:
            continue
        cluster = [a]
        for b in ungrouped[i + 1:]:
            if b in used:
                continue
            if min(len(a), len(b)) >= 3 and (a in b or b in a or SequenceMatcher(None, a, b).ratio() >= 0.86):
                cluster.append(b)
        if len(cluster) >= 2:
            label = min(cluster, key=len)
            for n in cluster:
                suggested[n] = label
                used.add(n)
    groups = defaultdict(list)
    for name in by_item:
        groups[suggested.get(name) or '未归组'].append(name)

    out = OUT_DIR / '2026Q4_按检测项目列抽取_同类项待判断.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = '同类项待判断'
    ws.append(['建议组ID', '建议统一检测项', '组内项目数', '组内记录数', '检测项目', '项目记录数', '示例检测结果', '示例单项判定', '示例PDF报告号', '示例货号', '示例颜色', 'PDF链接', '人工判断', '确认后统一名称', '备注'])
    gid = 0
    for label, names in sorted(groups.items(), key=lambda kv: -sum(len(by_item[n]) for n in kv[1])):
        if label == '未归组':
            continue
        gid += 1
        total = sum(len(by_item[n]) for n in names)
        group_id = f'G{gid:03d}'
        for idx, name in enumerate(sorted(names, key=lambda n: -len(by_item[n]))):
            rs = by_item[name]
            ex = rs[0]
            ws.append([group_id if idx == 0 else '', label if idx == 0 else '', len(names) if idx == 0 else '', total if idx == 0 else '', name, len(rs), ex.get('result',''), ex.get('verdict_raw',''), ex.get('report_no',''), ex.get('sku',''), ex.get('color',''), ex.get('url',''), '', '', ''])

    ws2 = wb.create_sheet('明细-检测项目列')
    ws2.append(['检测项目', '原始检测项目单元格', '检测项目细则', '检测结果', '检测结果细则', '单项判定', '判定归一', 'PDF报告号', '货号', '颜色', '源报告单号', 'PDF页码', 'PDF链接'])
    for r in sorted(rows, key=lambda x: (x.get('item',''), x.get('report_no',''), x.get('sku',''), x.get('color',''))):
        ws2.append([r.get('item',''), r.get('raw_item',''), r.get('item_detail',''), r.get('result',''), r.get('result_detail',''), r.get('verdict_raw',''), r.get('verdict',''), r.get('report_no',''), r.get('sku',''), r.get('color',''), r.get('source_order_no',''), r.get('page',''), r.get('url','')])

    ws3 = wb.create_sheet('未归组项目')
    ws3.append(['检测项目', '项目记录数', '示例检测结果', '示例单项判定', '示例PDF报告号', '示例货号', '示例颜色', 'PDF链接', '人工判断', '确认后统一名称', '备注'])
    for name in sorted(groups.get('未归组', []), key=lambda n: -len(by_item[n])):
        ex = by_item[name][0]
        ws3.append([name, len(by_item[name]), ex.get('result',''), ex.get('verdict_raw',''), ex.get('report_no',''), ex.get('sku',''), ex.get('color',''), ex.get('url',''), '', '', ''])

    for sheet in wb.worksheets:
        style_sheet(sheet)
    wb.save(out)

    state = json.loads(STATE_PATH.read_text(encoding='utf-8')) if STATE_PATH.exists() else {}
    summary = {
        'output': str(out),
        'processed_total': len(state.get('processed_urls', [])),
        'manifest_total': len(json.loads((WORK / 'download_manifest.json').read_text(encoding='utf-8'))),
        'extracted_rows': len(rows),
        'unique_items': len(by_item),
        'suggested_groups': gid,
        'ungrouped_items': len(groups.get('未归组', [])),
        'top_items': Counter({k: len(v) for k, v in by_item.items()}).most_common(30),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

if __name__ == '__main__':
    main()
