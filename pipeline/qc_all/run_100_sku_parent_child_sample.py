import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import extract_table_items_checkpoint as extractor


ROOT = Path.cwd()
WORK = ROOT / "tmp" / "qc_all"
OUT_DIR = ROOT / "outputs" / "019f4c2c-e544-7db3-a07e-ee8db394fef1"
OUT_PATH = OUT_DIR / "2026Q4_新逻辑_100款效果验证_V2.xlsx"

source = json.loads((WORK / "source_records.json").read_text(encoding="utf-8"))
manifest = json.loads((WORK / "download_manifest.json").read_text(encoding="utf-8"))
manifest_by_url = {row["url"]: row for row in manifest}

sample_skus = source["selected_skus"][:100]
sample_sku_set = set(sample_skus)

sample_records = [record for record in source["selected_records"] if record.get("sku") in sample_sku_set]
sample_urls = []
for record in sample_records:
    for url in record.get("urls", []):
        if url not in sample_urls:
            sample_urls.append(url)

detail_rows = []
stats_total = Counter()
for index, url in enumerate(sample_urls, 1):
    manifest_row = manifest_by_url.get(url)
    if not manifest_row:
        stats_total["missing_manifest"] += 1
        continue
    rows, stats = extractor.process_pdf(manifest_row)
    stats_total.update(stats)
    for row in rows:
        if row.get("sku") in sample_sku_set:
            detail_rows.append(row)
    if index % 50 == 0 or index == len(sample_urls):
        print(
            f"progress pdf={index}/{len(sample_urls)} detail_rows={len(detail_rows)}",
            flush=True,
        )

by_item = defaultdict(list)
for row in detail_rows:
    by_item[row["item"]].append(row)

short_terms = {"沾色", "变色", "干摩", "湿摩"}
short_term_rows = [row for row in detail_rows if row["item"] in short_terms]
parent_child_rows = [row for row in detail_rows if "-" in row["item"]]

wb = Workbook()
ws = wb.active
ws.title = "100款检测明细"
headers = [
    "货号",
    "颜色",
    "检测项目",
    "原始检测项目单元格",
    "检测项目细则",
    "检测结果",
    "检测结果细则",
    "单项判定",
    "PDF报告号",
    "源报告单号",
    "PDF页码",
    "PDF链接",
]
ws.append(headers)
for row in sorted(detail_rows, key=lambda r: (r.get("sku", ""), r.get("color", ""), r.get("item", ""), r.get("report_no", ""))):
    ws.append([
        row.get("sku", ""),
        row.get("color", ""),
        row.get("item", ""),
        row.get("raw_item", ""),
        row.get("item_detail", ""),
        row.get("result", ""),
        row.get("result_detail", ""),
        row.get("verdict_raw", ""),
        row.get("report_no", ""),
        row.get("source_order_no", ""),
        row.get("page", ""),
        row.get("url", ""),
    ])

ws2 = wb.create_sheet("检测项目汇总")
ws2.append(["检测项目", "记录数", "示例检测结果", "示例单项判定", "示例PDF报告号", "示例货号", "示例颜色", "PDF链接"])
for item, rows in sorted(by_item.items(), key=lambda pair: (-len(pair[1]), pair[0])):
    ex = rows[0]
    ws2.append([
        item,
        len(rows),
        ex.get("result", ""),
        ex.get("verdict_raw", ""),
        ex.get("report_no", ""),
        ex.get("sku", ""),
        ex.get("color", ""),
        ex.get("url", ""),
    ])

ws3 = wb.create_sheet("质量检查")
ws3.append(["检查项", "结果"])
ws3.append(["样本款号数", len(sample_skus)])
ws3.append(["样本源记录数", len(sample_records)])
ws3.append(["样本PDF数", len(sample_urls)])
ws3.append(["抽取明细行数", len(detail_rows)])
ws3.append(["去重检测项目数", len(by_item)])
ws3.append(["父级+细项行数", len(parent_child_rows)])
ws3.append(["短词单独成项行数", len(short_term_rows)])
ws3.append(["抽取统计", json.dumps(dict(stats_total), ensure_ascii=False)])

ws4 = wb.create_sheet("短词检查")
ws4.append(headers)
for row in short_term_rows:
    ws4.append([
        row.get("sku", ""),
        row.get("color", ""),
        row.get("item", ""),
        row.get("raw_item", ""),
        row.get("item_detail", ""),
        row.get("result", ""),
        row.get("result_detail", ""),
        row.get("verdict_raw", ""),
        row.get("report_no", ""),
        row.get("source_order_no", ""),
        row.get("page", ""),
        row.get("url", ""),
    ])

fill = PatternFill("solid", fgColor="1E6F7A")
thin = Side(style="thin", color="DCE5E8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for sheet in wb.worksheets:
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

widths = {
    "100款检测明细": [14, 18, 34, 36, 24, 32, 32, 12, 22, 18, 10, 70],
    "检测项目汇总": [34, 10, 32, 12, 22, 14, 18, 70],
    "质量检查": [24, 80],
    "短词检查": [14, 18, 34, 36, 24, 32, 32, 12, 22, 18, 10, 70],
}
for sheet in wb.worksheets:
    for index, width in enumerate(widths.get(sheet.title, []), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

OUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)

print(json.dumps({
    "output": str(OUT_PATH),
    "sample_skus": len(sample_skus),
    "sample_records": len(sample_records),
    "sample_pdfs": len(sample_urls),
    "detail_rows": len(detail_rows),
    "unique_items": len(by_item),
    "parent_child_rows": len(parent_child_rows),
    "short_term_rows": len(short_term_rows),
    "stats": dict(stats_total),
}, ensure_ascii=False, indent=2))
