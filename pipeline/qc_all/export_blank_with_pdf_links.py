import json
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

root = Path.cwd()
input_path = root/'outputs'/'019f4c2c-e544-7db3-a07e-ee8db394fef1'/'2026Q4_新增候选检测项_语义归并待确认.xlsx'
# 如果原表后续被删除，可改为从当前已导出的待判断表读取；这里优先读取原用户确认表。
fallback_path = root/'outputs'/'019f4c2c-e544-7db3-a07e-ee8db394fef1'/'2026Q4_新增候选检测项_待再次判断_附出处.xlsx'
out_path = root/'outputs'/'019f4c2c-e544-7db3-a07e-ee8db394fef1'/'2026Q4_新增候选检测项_待再次判断_附PDF链接.xlsx'
items = json.loads((root/'pipeline/qc_all/item_candidates.json').read_text(encoding='utf-8'))
item_by_name = {x['simple_item']: x for x in items}

if input_path.exists():
    wb = load_workbook(input_path, data_only=True)
    ws = wb['建议归并组']
    headers = [c.value for c in ws[1]]
    idx = {h:i for i,h in enumerate(headers)}
    current = {'gid':'', 'suggested':'', 'reason':''}
    blank = []
    rename_values = {'同意','排除','不同意','拆分',''}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx['归并组ID']]: current['gid'] = str(row[idx['归并组ID']]).strip()
        if row[idx['建议统一检测项']]: current['suggested'] = str(row[idx['建议统一检测项']]).strip()
        if row[idx['归并理由']]: current['reason'] = str(row[idx['归并理由']]).strip()
        candidate = str(row[idx['候选检测项']] or '').strip()
        if not candidate:
            continue
        judgment = str(row[idx['人工判断']] or '').strip()
        confirm = str(row[idx['确认后统一名称']] or '').strip()
        is_confirmed = judgment == '同意' or confirm or (judgment and judgment not in rename_values)
        if not judgment and not confirm and not is_confirmed:
            blank.append({'gid': current['gid'], 'suggested': current['suggested'], 'candidate': candidate,
                          'count': row[idx['出现次数']] or 0, 'reason': current['reason']})
else:
    wb = load_workbook(fallback_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h:i for i,h in enumerate(headers)}
    blank = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        blank.append({'gid': row[idx['归并组ID']] or '', 'suggested': row[idx['建议统一检测项']] or '',
                      'candidate': row[idx['候选检测项']] or '', 'count': row[idx['出现次数']] or 0,
                      'reason': row[idx['归并理由']] or ''})

out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = '待再次判断'
out_headers = ['归并组ID','建议统一检测项','候选检测项','出现次数','PDF链接','PDF报告号','货号','颜色','源报告单号','示例详情','归并理由','人工判断','确认后统一名称','备注']
out_ws.append(out_headers)
for r in blank:
    item = item_by_name.get(r['candidate'], {})
    examples = item.get('examples') or []
    if not examples:
        examples = [{}]
    # 最多给3条PDF出处，避免同一候选项过长；第一条放主行，其余换行写在同一单元格。
    urls=[]; report_nos=[]; skus=[]; colors=[]; orders=[]; details=[]
    for ex in examples[:3]:
        if ex.get('url'): urls.append(ex.get('url'))
        if ex.get('report_no'): report_nos.append(ex.get('report_no'))
        if ex.get('sku'): skus.append(str(ex.get('sku')))
        if ex.get('color'): colors.append(str(ex.get('color')))
        if ex.get('source_order_no'): orders.append(str(ex.get('source_order_no')))
        if ex.get('detail'): details.append(str(ex.get('detail')))
    out_ws.append([
        r['gid'], r['suggested'], r['candidate'], r['count'],
        '\n'.join(dict.fromkeys(urls)), '\n'.join(dict.fromkeys(report_nos)), '\n'.join(dict.fromkeys(skus)),
        '\n'.join(dict.fromkeys(colors)), '\n'.join(dict.fromkeys(orders)), '\n---\n'.join(details), r['reason'], '', '', ''
    ])

fill = PatternFill('solid', fgColor='1E6F7A')
thin = Side(style='thin', color='DCE5E8')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in out_ws[1]:
    c.fill = fill
    c.font = Font(color='FFFFFF', bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
for row in out_ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = border
widths = [12,28,30,10,85,24,14,16,24,80,40,14,24,24]
for i,w in enumerate(widths,1):
    out_ws.column_dimensions[get_column_letter(i)].width = w
out_ws.freeze_panes='A2'
out_ws.auto_filter.ref=out_ws.dimensions
out_wb.save(out_path)

missing_url = 0
for row in out_ws.iter_rows(min_row=2, values_only=True):
    if not row[4]:
        missing_url += 1
print(json.dumps({'output': str(out_path), 'rows': len(blank), 'missing_pdf_link_rows': missing_url}, ensure_ascii=False, indent=2))
