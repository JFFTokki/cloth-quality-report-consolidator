import json
import os
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "质检报告汇总.xlsx"
OUTPUT = Path(os.environ.get("QC_WORK_DIR", Path(__file__).resolve().parent)).resolve() / "source_records.json"
UNSPECIFIED_COLOR = "未标明颜色"
VALID_SEASON = os.environ.get("QC_SEASON", "2026Q4")
SELECTION_MODE = os.environ.get("QC_SOURCE_SELECTION", "all")
INVALID_SUBCATEGORIES = {"", "#N/A", "小类"}
EXCLUDE_SKUS_FILES = [
    value
    for name in ("QC_EXCLUDE_SKUS_FILE", "QC_EXCLUDE_SKUS_FILES")
    for value in os.environ.get(name, "").split(os.pathsep)
    if value.strip()
]


HEADER_ALIASES = {
    "order_no": ("ORDER_NO", "报告单号", "源报告单号"),
    "sku": ("PRD_CODE", "货号", "款号"),
    "subcategory": ("小类",),
    "color_raw": ("PRD_F1_DISPLAY", "颜色"),
    "material_code": ("P_MAT_CODE", "物料编码", "面料编码"),
    "material_color": ("MAT_F1_DISPLAY", "物料颜色", "面料颜色"),
    "brand": ("BRAND_DISPLAY", "品牌"),
    "season": ("SEASONS", "季度"),
    "sample_type": ("THIRDPARTY_TESTING_TESTPART_DISPLAY", "样品类型", "检测部位"),
    "overall_result": ("RESULT", "总体判定", "判定"),
    "attachment_raw": ("QC_ATTACHMENT_LIST", "附件", "PDF链接"),
    "modified_by": ("LAST_MODIFIED_BY", "最后修改人"),
    "modified_time": ("LAST_MODIFIED_TIME", "最后修改时间"),
}


def split_colors(value: str) -> list[str]:
    colors = [part.strip() for part in re.split(r"[,，;；\n]+", str(value or "")) if part.strip()]
    return colors or [UNSPECIFIED_COLOR]


def split_urls(value: str) -> list[str]:
    urls = re.findall(r"https?://[^;；,，\s]+", str(value or ""))
    return list(dict.fromkeys(url.rstrip(".;；，,") for url in urls))


def invalid_sku(sku: str) -> bool:
    return not sku or sku == "0" or re.search(r"[,，;；\n]", sku)


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_column(headers: list[str], logical_name: str, required: bool = True) -> int | None:
    aliases = HEADER_ALIASES[logical_name]
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    if required:
        raise SystemExit(f"源表缺少必要列：{logical_name}，候选表头={aliases}")
    return None


def read_source_rows() -> tuple[list[dict], list[dict]]:
    wb = load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    source_sheet = ws.title
    rows_iter = ws.iter_rows(values_only=True)
    headers = [cell_text(value) for value in next(rows_iter)]
    columns = {
        key: find_column(headers, key, required=key != "subcategory" or SELECTION_MODE == "one_per_subcategory")
        for key in HEADER_ALIASES
    }

    records = []
    invalid_records = []
    for source_row, row in enumerate(rows_iter, start=2):
        values = {
            key: cell_text(row[index]) if index is not None and index < len(row) else ""
            for key, index in columns.items()
        }
        if values["order_no"] == "ORDER_NO" and values["sku"] == "PRD_CODE":
            continue
        if VALID_SEASON.lower() != "all" and values["season"] != VALID_SEASON:
            continue

        sku = values["sku"]
        urls = split_urls(values["attachment_raw"])
        record = {
            "source_sheet": source_sheet,
            "source_row": source_row,
            "source_cell": f"{get_column_letter(columns['attachment_raw'] + 1)}{source_row}" if columns["attachment_raw"] is not None else "",
            "order_no": values["order_no"],
            "sku": sku,
            "subcategory": values.get("subcategory", ""),
            "color_raw": values["color_raw"],
            "colors": split_colors(values["color_raw"]),
            "material_code": values["material_code"],
            "material_color": values["material_color"],
            "brand": values["brand"],
            "season": values["season"],
            "sample_type": values["sample_type"],
            "overall_result": values["overall_result"],
            "attachment_raw": values["attachment_raw"],
            "urls": urls,
            "modified_by": values["modified_by"],
            "modified_time": values["modified_time"],
            "invalid_reason": "",
            "is_selected": False,
            "selected_colors": [],
        }
        if invalid_sku(sku):
            record["invalid_reason"] = "无有效货号"
            invalid_records.append(record)
        elif not urls:
            record["invalid_reason"] = "无PDF链接"
        records.append(record)
    wb.close()
    return records, invalid_records


def load_excluded_skus(paths: list[str]) -> set[str]:
    excluded = set()
    for filename in paths:
        payload = json.loads(Path(filename).read_text(encoding="utf-8"))
        if isinstance(payload, list):
            values = payload
        elif isinstance(payload, dict):
            values = payload.get("skus") or payload.get("selected_skus") or []
        else:
            values = []
        for value in values:
            excluded.update(re.findall(r"(?<!\d)\d{12}(?!\d)", str(value)))
    return excluded


def choose_one_sku_per_subcategory(records: list[dict], excluded_skus: set[str]) -> tuple[list[str], list[dict]]:
    grouped = defaultdict(list)
    for record in records:
        subcategory = cell_text(record.get("subcategory"))
        if subcategory in INVALID_SUBCATEGORIES:
            continue
        if invalid_sku(record.get("sku", "")) or not record.get("urls"):
            continue
        grouped[subcategory].append(record)

    selected_skus = []
    used_skus = set()
    selections = []
    for subcategory in sorted(grouped):
        candidates = grouped[subcategory]
        preferred = [
            row for row in candidates
            if row["sku"] not in used_skus and row["sku"] not in excluded_skus
        ]
        chosen = next(iter(preferred), None)
        if chosen is None:
            chosen = next((row for row in candidates if row["sku"] not in used_skus), candidates[0])
        selected_skus.append(chosen["sku"])
        used_skus.add(chosen["sku"])
        candidate_skus = {row["sku"] for row in candidates}
        selections.append({
            "subcategory": subcategory,
            "sku": chosen["sku"],
            "source_row": chosen["source_row"],
            "order_no": chosen["order_no"],
            "candidate_records": len(candidates),
            "candidate_skus": len(candidate_skus),
            "excluded_candidate_skus": len(candidate_skus & excluded_skus),
            "used_excluded_fallback": chosen["sku"] in excluded_skus,
        })
    return selected_skus, selections


records, invalid_records = read_source_rows()
excluded_skus = load_excluded_skus(EXCLUDE_SKUS_FILES)

if SELECTION_MODE == "one_per_subcategory":
    selected_skus, subcategory_selections = choose_one_sku_per_subcategory(records, excluded_skus)
    selected_sku_set = set(selected_skus)
else:
    subcategory_selections = []
    selected_sku_set = {
        record["sku"]
        for record in records
        if not invalid_sku(record.get("sku", "")) and record.get("urls")
    }
    selected_skus = sorted(selected_sku_set)

for record in records:
    if not record.get("urls") or invalid_sku(record.get("sku", "")):
        continue
    if record["sku"] in selected_sku_set:
        record["is_selected"] = True
        record["selected_colors"] = record["colors"]

selected_records = [record for record in records if record["is_selected"]]
selected_urls = list(dict.fromkeys(url for record in selected_records for url in record["urls"]))
summary_keys = sorted(
    [
        {"sku": record["sku"], "color": color}
        for record in selected_records
        for color in record["selected_colors"]
    ],
    key=lambda row: (row["sku"], row["color"]),
)
seen = set()
summary_keys = [
    row for row in summary_keys
    if (row["sku"], row["color"]) not in seen and not seen.add((row["sku"], row["color"]))
]

payload = {
    "source_workbook": str(INPUT),
    "season": VALID_SEASON,
    "selection_mode": SELECTION_MODE,
    "exclude_skus_files": EXCLUDE_SKUS_FILES,
    "excluded_skus_count": len(excluded_skus),
    "selected_skus": selected_skus,
    "selected_subcategories": [row["subcategory"] for row in subcategory_selections],
    "subcategory_selections": subcategory_selections,
    "records": records,
    "invalid_records": invalid_records,
    "selected_records": selected_records,
    "selected_urls": selected_urls,
    "summary_keys": summary_keys,
}
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(
    json.dumps(
        {
            "season": VALID_SEASON,
            "selection_mode": SELECTION_MODE,
            "selected_skus": len(payload["selected_skus"]),
            "selected_subcategories": len(payload["selected_subcategories"]),
            "excluded_skus_count": len(excluded_skus),
            "excluded_fallback_subcategories": sum(
                1 for row in subcategory_selections if row.get("used_excluded_fallback")
            ),
            "source_records": len(records),
            "invalid_records": len(invalid_records),
            "selected_source_records": len(selected_records),
            "summary_rows": len(summary_keys),
            "unique_selected_urls": len(selected_urls),
            "multi_url_selected_records": sum(len(record["urls"]) > 1 for record in selected_records),
            "unspecified_color_records": sum(UNSPECIFIED_COLOR in record["colors"] for record in selected_records),
        },
        ensure_ascii=False,
        indent=2,
    )
)
