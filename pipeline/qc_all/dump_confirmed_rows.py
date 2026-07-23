from pathlib import Path
from openpyxl import load_workbook
import json
p=Path('outputs/manual_review/2026Q4_新增候选检测项_语义归并待确认.xlsx')
wb=load_workbook(p,data_only=True)
ws=wb['建议归并组']
headers=[c.value for c in ws[1]]
idx={h:i for i,h in enumerate(headers)}
current={}
rows=[]
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[idx['归并组ID']]: current['gid']=r[idx['归并组ID']]
    if r[idx['建议统一检测项']]: current['suggested']=r[idx['建议统一检测项']]
    judgment=str(r[idx['人工判断']] or '').strip()
    confirm=str(r[idx['确认后统一名称']] or '').strip()
    if judgment or confirm:
        rows.append({
            'gid': current.get('gid',''), 'suggested': current.get('suggested',''),
            'candidate': r[idx['候选检测项']], 'count': r[idx['出现次数']],
            'judgment': judgment, 'confirm': confirm, 'detail': r[idx['示例详情']]
        })
print(json.dumps(rows, ensure_ascii=False, indent=2))
