from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

path = Path('outputs/manual_review/2026Q4_新增候选检测项_语义归并待确认.xlsx')
wb = load_workbook(path, data_only=True)
for ws in wb.worksheets:
    headers=[c.value for c in ws[1]]
    print('\nSHEET', ws.title)
    print(headers)
    if '人工判断' in headers:
        i=headers.index('人工判断')+1
        j=headers.index('确认后统一名称')+1 if '确认后统一名称' in headers else None
        c=Counter()
        renamed=0
        for row in ws.iter_rows(min_row=2, values_only=True):
            c[str(row[i-1]).strip() if row[i-1] is not None else ''] += 1
            if j and row[j-1] not in (None,''):
                renamed += 1
        print('人工判断分布', dict(c))
        print('确认后统一名称非空', renamed)
