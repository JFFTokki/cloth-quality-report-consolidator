import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from report_header import normalize_institution_display_value


ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parents[1]
source = json.loads((ROOT / "source_records.json").read_text(encoding="utf-8"))
documents = json.loads((ROOT / "pdf_text.json").read_text(encoding="utf-8"))
document_by_url = {document["url"]: document for document in documents}
UNSPECIFIED_COLOR = "未标明颜色"

PHRASE_REPLACEMENTS = {
    "耐汗漬": "耐汗渍", "洗後外觀": "洗后外观", "脹破強度": "胀破强度",
    "耐光汗復合": "耐光汗复合", "儲存色遷移": "储存色迁移", "酚黃變": "酚黄变",
    "織物掉纖維": "织物掉纤维", "甲醛含量": "甲醛含量", "檢測": "检测",
    "報告": "报告", "結果": "结果", "標準": "标准", "實測": "实测",
    "纖維": "纤维", "錦綸": "锦纶", "腈綸": "腈纶", "聚酯纖維": "聚酯纤维",
    "變色": "变色", "沾色": "沾色", "濕摩": "湿摩", "堿": "碱",
    "未檢出": "未检出", "無": "无", "符合": "符合",
}

RULES = [
    ("纤维含量", ["纤维含量", "纤维成分及含量", "纤维成分"]),
    ("pH值", ["pH值（其他）", "pH值(其他)", "pH值", "PH值"]),
    ("甲醛含量", ["甲醛含量", "甲醛"]),
    ("可分解致癌芳香胺染料", ["可分解致癌芳香胺染料", "可分解致癌芳香胺", "禁用偶氮染料", "香胺染料"]),
    ("异味", ["异味"]),
    ("耐光汗复合色牢度", ["耐光汗复合色牢度", "耐光汗复合色牢"]),
    ("耐光色牢度", ["耐光色牢度"]),
    ("耐贮存色牢度", ["耐贮存色牢度", "储存色迁移"]),
    ("酚黄变色牢度", ["酚黄变色牢度", "酚黄变"]),
    ("耐皂洗色牢度", ["耐皂洗色牢度", "耐洗色牢度", "皂洗色牢度", "耐皂洗", "耐洗", "GB/T 3921", "GB/T3921"]),
    ("洗液沾色程度", ["洗液沾色程度"]),
    ("耐汗渍色牢度", ["耐酸汗渍色牢度", "耐碱汗渍色牢度", "耐汗渍色牢度", "酸汗", "碱汗", "汗渍色牢度", "耐汗渍"]),
    ("耐水色牢度", ["耐水色牢度", "耐水", "GB/T 5713", "GB/T5713"]),
    ("耐摩擦色牢度", ["耐干摩擦色牢度", "耐湿摩擦色牢度", "耐摩擦色牢度", "干摩擦色牢度", "湿摩擦色牢度", "干摩擦", "湿摩擦", "干摩", "湿摩"]),
    ("耐唾液色牢度", ["耐唾液色牢度", "耐唾液"]),
    ("拼接互染", ["拼接互染色牢度", "拼接互染"]),
    ("洗后外观", ["水洗后外观", "洗后外观"]),
    ("水洗尺寸变化率", ["水洗尺寸变化率"]),
    ("抗起毛球", ["抗起毛球", "面料起球", "起球"]),
    ("胀破强度", ["胀破强度"]),
    ("断裂强力", ["断裂强力"]),
    ("撕破强力", ["撕破强力"]),
    ("缝子纰裂程度", ["缝子纰裂程度", "纰裂程度"]),
    ("附件抗拉强力", ["附件抗拉强力", "附件抗拉强度"]),
    ("锐利尖端和锐利边缘", ["锐利尖端和锐利边缘", "附件锐利性"]),
    ("绳带要求", ["绳带要求", "绳带"]),
    ("可萃取重金属", ["可萃取重金属"]),
    ("邻苯二甲酸酯", ["邻苯二甲酸酯"]),
    ("重金属总量", ["重金属总量"]),
    ("富马酸二甲酯", ["富马酸二甲酯"]),
    ("含氯苯酚", ["含氯苯酚"]),
    ("短链氯化石蜡", ["短链氯化石蜡"]),
    ("蓬松度", ["蓬松度"]),
    ("绒子含量", ["绒子含量"]),
    ("充绒量", ["充绒量"]),
    ("钻绒值", ["钻绒值"]),
    ("掉纤维程度", ["织物掉纤维", "掉纤维程度"]),
    ("使用说明", ["使用说明"]),
    ("标识标签", ["标识标签*", "标识标签", "标识"]),
    ("感官质量", ["感官质量"]),
    ("衬里和内垫耐摩擦色牢度", ["衬里和内垫耐摩擦色牢度"]),
    ("底墙与帮面剥离强度", ["底墙与帮面剥离强度"]),
    ("外底耐磨性能", ["外底耐磨性能"]),
    ("成鞋耐折性能", ["成鞋耐折性能"]),
    ("婴幼儿鞋小附件要求", ["婴幼儿鞋小附件要求"]),
    ("可分解致癌芳香胺染料-具体芳香胺物质", ["二氯联苯胺"]),
    ("外观面料", ["外观面料"]),
    ("延伸值", ["横向延伸值"]),
    ("异色毛绒", ["异色毛绒"]),
    ("抗静电性能", ["抗静电性能"]),
    ("掉纤测试", ["掉纤测试"]),
    ("杂质含量", ["杂质"]),
    ("残留金属针", ["残留金属针"]),
    ("汞Hg", ["汞Hg"]),
    ("砷As", ["砷As"]),
    ("绒丝+羽丝含量", ["绒丝羽丝", "绒丝+羽丝"]),
    ("维护方法", ["维护方法"]),
    ("耐水洗性", ["耐水洗性"]),
    ("耐腐蚀性", ["五金配件耐腐蚀性"]),
    ("聚酯纤维", ["聚酯纤维"]),
    ("脱毛量mg", ["脱毛量mg", "脱毛量"]),
    ("透湿性能", ["透湿性能"]),
    ("烷基酚和烷基酚聚氧乙烯醚", ["酚聚氧乙烯醚", "烷基酚和烷基酚聚氧乙烯醚"]),
    ("里料不允许外露", ["里料不允许外露"]),
    ("重金属", ["可萃取汞", "重金属"]),
    ("钴Co", ["钴Co"]),
    ("铅Pb", ["铅Pb"]),
    ("铜Cu", ["铜Cu"]),
    ("铬Cr", ["铬Cr"]),
    ("锑Sb", ["锑Sb"]),
    ("镍Ni", ["镍Ni"]),
    ("防污性能", ["防污性能"]),
    ("防油性能", ["防油性能"]),
    ("防泼水性能", ["防泼水性能"]),
    ("防钻绒性", ["防钻绒性"]),
    ("防风性能", ["防风性能"]),
    ("附件尖端和边缘锐利性", ["附件尖端和边缘的"]),
    ("附件锐利性", ["金属附件允许轻微掉漆"]),
    ("陆禽毛含量", ["陆禽毛含量"]),
    ("顶破强力", ["顶破强力"]),
    ("（绒丝+羽丝）含量", ["含量森马集团羽绒采购内控指标"]),
]

ORDER = [name for name, _ in RULES]
ALL_ALIASES = [(standard, alias) for standard, aliases in RULES for alias in aliases]
REVIEW_RULES = []


def load_review_rules():
    path = PROJECT_ROOT / "outputs" / "manual_review" / "2026Q4_按检测项目列抽取_同类项待判断.xlsx"
    if not path.exists():
        return []

    output = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in ("同类项待判断", "未归组项目"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "检测项目" not in headers or "确认后统一名称" not in headers:
            continue
        item_i = headers.index("检测项目")
        confirm_i = headers.index("确认后统一名称")
        judgment_i = headers.index("人工判断") if "人工判断" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = str(row[item_i] or "").strip()
            confirmed = str(row[confirm_i] or "").strip()
            judgment = str(row[judgment_i] or "").strip() if judgment_i is not None else ""
            if not item or not confirmed or judgment == "排除":
                continue
            output.append((confirmed, item))
    return output


REVIEW_RULES = load_review_rules()
for standard, alias in REVIEW_RULES:
    RULES.append((standard, [alias]))
    if standard not in ORDER:
        ORDER.append(standard)
    ALL_ALIASES.append((standard, alias))


def simplified(value: str) -> str:
    output = value or ""
    for old, new in PHRASE_REPLACEMENTS.items():
        output = output.replace(old, new)
    return output


def normalize_line(value: str) -> str:
    return re.sub(r"\s+", " ", simplified(value)).strip()


def verdict_from_line(line: str) -> str:
    if re.search(r"不符合|不合格", line):
        return "不符合"
    if re.search(r"符合|合格|通过", line):
        return "符合"
    if "实测" in line:
        return "实测"
    return ""


def result_from_line(line: str, verdict: str, standard_item: str) -> str:
    before = re.split(r"不符合|不合格|符合|合格|通过", line, maxsplit=1)[0]
    before = re.sub(r"^\s*\d{1,3}[.)、]?\s+", "", before)
    if re.search(r"未检出|N\.?D\.?", before, re.I):
        return "未检出"
    for phrase in ("无异味", "无缺陷", "没有此项", "无明显变化"):
        if phrase in before:
            return phrase
    has_threshold = bool(re.search(r"[≥≤≧≦<>～~]", before))
    cleaned = re.sub(r"(?:GB|FZ|QB|Q|ISO|AATCC|ASTM)[/\sA-Z.-]*\d[\d.:-]*", " ", before, flags=re.I)
    tokens = re.findall(r"(?<![A-Za-z])[-+]?\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?(?:%|cm|mm|mg/kg|N|g|级|根/m²)?", cleaned)
    usable = []
    for token in tokens:
        plain = re.sub(r"[^\d.]", "", token)
        if plain and plain.isdigit() and 1900 <= int(plain) <= 2100:
            continue
        if len(plain) >= 7:
            continue
        usable.append(token)
    if not usable:
        return ""
    if has_threshold:
        return usable[-1] if len(usable) >= 2 else ""
    direct_value_items = {"pH值", "蓬松度", "绒子含量", "充绒量", "钻绒值", "胀破强度", "抗起毛球", "掉纤维程度"}
    if standard_item in direct_value_items and len(usable) == 1 and not re.search(r"标准|方法|温度|时间|次数|℃|V\d", before, re.I):
        return usable[0]
    return ""


def better_institution(document: dict) -> str:
    current = normalize_institution_display_value(simplified(document.get("institution", "")))
    if current != "未识别" and not any(token in current for token in ("日期", "类别", "性质", "序号", "委托检验", "检测报告", "检验报告", "供应商", "委托方", "送样人", "浙江森马")):
        return current
    first_text = "\n".join(page["text"] for page in document.get("pages", [])[:2])
    lines = [normalize_line(line) for line in first_text.splitlines() if line.strip()]
    for line in lines[:80]:
        company_match = re.search(r"([\u4e00-\u9fff（）()]{4,}(?:有限公司|研究院|检测中心|检验中心|质量中心))", line)
        if (
            company_match
            and any(token in company_match.group(1) for token in ("检测", "检验", "质量", "计量", "纺织"))
            and not any(token in line for token in ("委托单位", "委托方", "供应商", "生产单位", "制造单位", "送样人", "地址", "电话", "浙江森马"))
        ):
            return company_match.group(1)
    return "未识别"


def report_no(document: dict) -> str:
    value = document.get("report_no") or Path(document["url"].split("?")[0]).stem[:40]
    if value.upper() in {"CATEGORY", "INSPECTION"} or len(value) < 5:
        return Path(document["url"].split("?")[0]).stem[:40]
    return value


def candidate_item_from_line(line: str) -> str:
    compact_line = re.sub(r"\s+", "", line).lower()
    if any(re.sub(r"\s+", "", alias).lower() in compact_line for _, alias in ALL_ALIASES):
        return ""
    if not verdict_from_line(line):
        return ""
    if not re.search(r"[\u4e00-\u9fff]", line):
        return ""
    noise_tokens = (
        "报告编号", "报告日期", "委托单位", "样品名称", "商标", "地址", "电话", "检测日期", "收样日期", "检验结论", "二维码",
        "检验类别", "检测类别", "委托送样", "委托检测", "到样日期", "样品状态", "样品等级", "产品等级", "安全类别",
        "序号", "检测方法", "标准值", "实测值", "单项判定", "单项评价", "报告检测结果", "不确定度", "符合性判定",
        "符合性声明", "规定限值", "检测结果与符合性", "年份/季节", "样品信息", "产品信息", "见下表", "见下列",
        "检验类别", "检测類別", "委託检测", "樣品狀態", "所检项目", "所檢項目", "本报告", "本報告", "检验检测结论",
        "检验结果", "检测结果", "实测结果", "定量限", "型号", "规格", "年份", "季节", "样品描述", "产品名称",
        "产品号型", "执行的产品标准", "友情提醒", "声明", "备注", "符合检验要求", "符合检测要求", "整體評價",
    )
    if any(token in line for token in noise_tokens):
        return ""
    if re.search(r"见下(?:\([^)]*\)|（[^）]*）)?表|见附表|详见.*表", line):
        return ""
    before = re.split(r"不符合|不合格|符合|合格|通过", line, maxsplit=1)[0]
    before = re.sub(r"^\s*\d{1,3}[.)、]?\s*", "", before)
    before = re.split(r"\b(?:GB|FZ|QB|ISO|AATCC|ASTM|SN|EN|JIS)[/\sA-Z.-]*\d", before, maxsplit=1, flags=re.I)[0]
    before = re.sub(r"\[[^\]]{1,20}\]|（[^）]{1,20}）|\([^)]{1,20}\)", "", before)
    before = re.sub(r"\b\d{2,7}-\d{2,7}-\d{1,7}\b", " ", before)
    before = re.split(r"客户要求|采购内控标准|Q/|Semir|[≤≥<>＝=]|N\.?D\.?|未检出|mg/kg|cm|mm|级", before, maxsplit=1, flags=re.I)[0]
    before = re.sub(r"[/／:：].*$", "", before).strip()
    before = re.sub(r"^[\d\s,，.、'\"-]+", "", before)
    before = re.sub(r"[\d\s,，.、'\"%_-]+$", "", before)
    before = re.sub(r"[^A-Za-z\u4e00-\u9fff]+", "", before)
    before = re.sub(r"(?:等级|级)$", "", before)
    if len(before) == 1 and before not in {"铅", "镉", "汞", "砷", "铬", "镍"}:
        return ""
    if not (1 <= len(before) <= 28):
        return ""
    if before in {"序号", "类别", "检测项目", "单项判定", "备注", "标准要求", "实测结果", "检验结论", "沾色", "变色", "色牢度"}:
        return ""
    if not re.search(r"[\u4e00-\u9fff]", before):
        return ""
    return before


def project_lines_only(lines):
    """Return lines from inspection/test item tables, not basis/statement sections."""
    output = []
    in_table = False
    remaining_after_header = 0
    stop_tokens = ("判定依据", "检验依据", "检测依据", "检测结论", "检验结论", "备注", "声明", "注意事项", "签发日期", "批准", "审核", "编制")
    header_re = re.compile(r"(序号\s*)?(检验|检测|测试)(项目|項目)|检验项目汇总|检测项目汇总|检测结果|檢測結果")
    single_item_re = re.compile(r"^(检验|检测|测试)(项目|項目)\s+(.{2,60})$")

    for page_number, line in lines:
        single = single_item_re.match(line)
        if single and not any(token in line for token in ("汇总", "见附页", "见数据页")):
            output.append((page_number, line))
            continue

        if header_re.search(line):
            in_table = True
            remaining_after_header = 120
            # Header rows define the table but are not themselves detection items.
            continue

        if any(token in line for token in stop_tokens):
            if not line.startswith(("检测项目", "检验项目", "测试项目")):
                in_table = False
                remaining_after_header = 0
                continue

        if not in_table:
            continue
        if re.search(r"^(第\s*\d+\s*页|报告编号|样品照片|以下空白|—以下空白|DECLARATION)", line, re.I):
            in_table = False
            remaining_after_header = 0
            continue
        if remaining_after_header <= 0:
            in_table = False
            continue

        output.append((page_number, line))
        remaining_after_header -= 1
    return output


def parse_document(document: dict) -> dict:
    institution = better_institution(document)
    current_report_no = report_no(document)
    doc_errors = []
    if document["status"] == "failed":
        doc_errors.append({"type": "PDF下载失败", "detail": document.get("error", "")})
        return {"report_no": "", "institution": institution, "items": [], "errors": doc_errors, "status": "下载失败"}
    if document.get("extract_error"):
        doc_errors.append({"type": "PDF解析失败", "detail": document["extract_error"]})
    if document.get("text_chars", 0) < 300:
        doc_errors.append({"type": "PDF文本不足", "detail": f'仅提取{document.get("text_chars", 0)}个字符，可能为扫描件'})
    if institution == "未识别":
        doc_errors.append({"type": "机构名称未识别", "detail": "报告文本没有可用的检测机构文字"})

    lines = []
    for page in document.get("pages", []):
        for raw_line in page.get("text", "").splitlines():
            line = normalize_line(raw_line)
            if line:
                lines.append((page["page"], line))
    project_lines = project_lines_only(lines)

    found = {}
    mapping_rows = {}
    for standard_item, aliases in RULES:
        matches = []
        for page_number, line in project_lines:
            matched_alias = next((alias for alias in aliases if alias.lower() in line.lower()), None)
            if not matched_alias:
                continue
            verdict = verdict_from_line(line)
            if not verdict:
                continue
            result = result_from_line(line, verdict, standard_item)
            score = (5 if result else 0) + (2 if re.match(r"^\d+", line) else 0) - len(line) / 500
            matches.append((score, page_number, line, matched_alias, verdict, result))
        if not matches:
            continue
        matches.sort(reverse=True)
        _, page_number, evidence, raw_item, verdict, result = matches[0]
        found[standard_item] = {
            "raw_item": raw_item,
            "simple_item": simplified(raw_item),
            "standard_item": standard_item,
            "subitem": "",
            "result": result,
            "unit": "",
            "requirement": "",
            "method": "",
            "verdict": verdict,
            "page": page_number,
            "evidence": evidence,
            "status": "已确认",
        }
        mapping_rows[(institution, raw_item, standard_item)] = {
            "institution": institution,
            "raw_item": raw_item,
            "simple_item": simplified(raw_item),
            "standard_item": standard_item,
            "merge_basis": "机构原名即统一检测项名称" if simplified(raw_item) == standard_item else f'不同机构命名差异，统一归为“{standard_item}”',
            "status": "已确认",
            "first_report_no": current_report_no,
            "first_url": document["url"],
        }

    candidate_counter = Counter()
    candidate_examples = {}
    for page_number, line in project_lines:
        candidate = candidate_item_from_line(line)
        if not candidate:
            continue
        candidate_counter[candidate] += 1
        candidate_examples.setdefault(candidate, (page_number, line))

    for candidate, count in candidate_counter.items():
        if candidate in found:
            continue
        page_number, evidence = candidate_examples[candidate]
        verdict = verdict_from_line(evidence)
        result = result_from_line(evidence, verdict, candidate)
        found[candidate] = {
            "raw_item": candidate,
            "simple_item": simplified(candidate),
            "standard_item": simplified(candidate),
            "subitem": "",
            "result": result,
            "unit": "",
            "requirement": "",
            "method": "",
            "verdict": verdict,
            "page": page_number,
            "evidence": evidence,
            "status": "待确认",
        }
        mapping_rows[(institution, candidate, simplified(candidate))] = {
            "institution": institution,
            "raw_item": candidate,
            "simple_item": simplified(candidate),
            "standard_item": simplified(candidate),
            "merge_basis": "新增检测项，需人工确认是否新增标准项或归并为既有项目",
            "status": "新增候选",
            "first_report_no": current_report_no,
            "first_url": document["url"],
        }
        doc_errors.append({
            "type": "新增检测项待确认",
            "raw_item": candidate,
            "simple_item": simplified(candidate),
            "suggested_item": simplified(candidate),
            "detail": f"出现{count}次；示例：{evidence[:160]}",
        })

    if not found:
        doc_errors.append({"type": "未识别检测项", "detail": "报告有文本，但未匹配到任何已定义或候选检测项"})
        status = "PDF文本不足" if document.get("text_chars", 0) < 300 else "未识别检测项"
    elif any(row.get("status") == "待确认" for row in found.values()):
        status = "新增检测项待确认"
    elif institution == "未识别":
        status = "机构未识别但已解析"
    else:
        status = "已解析"

    return {
        "report_no": current_report_no,
        "institution": institution,
        "items": list(found.values()),
        "errors": doc_errors,
        "mapping_rows": list(mapping_rows.values()),
        "status": status,
    }


parsed_documents = {document["url"]: parse_document(document) for document in documents}
mapping = {}
for parsed in parsed_documents.values():
    for row in parsed.get("mapping_rows", []):
        mapping[(row["institution"], row["raw_item"], row["standard_item"])] = row

records_by_key = defaultdict(list)
for record in source["selected_records"]:
    for color in record["selected_colors"]:
        records_by_key[(record["sku"], color)].append(record)

summary_rows = []
detail_rows = []
final_errors = []
item_candidates = defaultdict(lambda: {"count": 0, "examples": []})

for key in source["summary_keys"]:
    sku, color = key["sku"], key["color"]
    records = records_by_key[(sku, color)]
    item_values = defaultdict(list)
    report_nos = []
    urls = []
    order_nos = []
    modified_times = []
    overall_results = []
    for record in records:
        order_nos.append(record["order_no"])
        modified_times.append(record["modified_time"])
        overall_results.append(record["overall_result"])
        for url in record["urls"]:
            urls.append(url)
            parsed = parsed_documents.get(url, {"items": [], "errors": [], "report_no": "", "institution": ""})
            if parsed.get("report_no"):
                report_nos.append(parsed["report_no"])
            for error in parsed.get("errors", []):
                contextual = {
                    "type": error.get("type", ""),
                    "sku": sku,
                    "color": color,
                    "source_order_no": record["order_no"],
                    "report_no": parsed.get("report_no", ""),
                    "url": url,
                    "raw_item": error.get("raw_item", ""),
                    "simple_item": error.get("simple_item", ""),
                    "suggested_item": error.get("suggested_item", ""),
                    "detail": error.get("detail", ""),
                }
                final_errors.append(contextual)
                if contextual["type"] == "新增检测项待确认":
                    item_candidates[contextual["simple_item"]]["count"] += 1
                    if len(item_candidates[contextual["simple_item"]]["examples"]) < 5:
                        item_candidates[contextual["simple_item"]]["examples"].append(contextual)
            for row in parsed.get("items", []):
                detail_rows.append({
                    "sku": sku,
                    "color": color,
                    "source_order_no": record["order_no"],
                    "report_no": parsed.get("report_no", ""),
                    "institution": parsed.get("institution", ""),
                    **row,
                    "url": url,
                })
                display = row["result"] if row["verdict"] == "符合" else (f'{row["result"]}｜{row["verdict"]}' if row["result"] else row["verdict"])
                if not display and row["verdict"] == "符合":
                    continue
                tagged = {"report_no": parsed.get("report_no", ""), "display": display}
                if tagged not in item_values[row["standard_item"]]:
                    item_values[row["standard_item"]].append(tagged)
    for standard_item, values in item_values.items():
        distinct = {value["display"] for value in values}
        if len(distinct) > 1:
            final_errors.append({
                "type": "同项结果差异",
                "sku": sku,
                "color": color,
                "source_order_no": "\n".join(dict.fromkeys(order_nos)),
                "report_no": " / ".join(value["report_no"] for value in values),
                "url": "\n".join(dict.fromkeys(urls)),
                "raw_item": "",
                "simple_item": standard_item,
                "suggested_item": standard_item,
                "detail": f'{sku} {color} {standard_item}：' + "；".join(f'{v["report_no"]}={v["display"]}' for v in values),
            })
    sorted_times = sorted(t for t in modified_times if t)
    summary_rows.append({
        "sku": sku,
        "color": color,
        "source_order_no": "\n".join(dict.fromkeys(order_nos)),
        "source_modified_time": f"{sorted_times[0]} ~ {sorted_times[-1]}" if sorted_times else "",
        "report_nos": "\n".join(dict.fromkeys(report_nos)),
        "overall_result": "\n".join(dict.fromkeys(overall_results)),
        "urls": "\n".join(dict.fromkeys(urls)),
        "items": dict(item_values),
    })

source_index = []
for record in source["records"] + source.get("invalid_records", []):
    urls = record["urls"] or [""]
    for url in urls:
        if record.get("invalid_reason"):
            status = record["invalid_reason"]
            selected = "否"
            reason = record["invalid_reason"]
            if status == "无有效货号":
                final_errors.append({
                    "type": "无有效货号",
                    "sku": record["sku"],
                    "color": record["color_raw"] or UNSPECIFIED_COLOR,
                    "source_order_no": record["order_no"],
                    "report_no": "",
                    "url": url,
                    "raw_item": "",
                    "simple_item": "",
                    "suggested_item": "",
                    "detail": f'源表第{record["source_row"]}行货号为空、为0或复合货号',
                })
        elif not url:
            status = "无PDF链接"
            selected = "否"
            reason = "源记录没有可提取的PDF链接"
            final_errors.append({
                "type": "无PDF链接",
                "sku": record["sku"],
                "color": record["color_raw"] or UNSPECIFIED_COLOR,
                "source_order_no": record["order_no"],
                "report_no": "",
                "url": "",
                "raw_item": "",
                "simple_item": "",
                "suggested_item": "",
                "detail": f'源表第{record["source_row"]}行没有可提取PDF链接',
            })
        else:
            parsed = parsed_documents.get(url, {"status": "未识别检测项", "errors": []})
            status = parsed["status"]
            selected = "是"
            reason = "；".join(dict.fromkeys(f'{e.get("type", "")}: {e.get("detail", "")}' for e in parsed.get("errors", [])))
        source_index.append({
            "source_row": record["source_row"],
            "sku": record["sku"],
            "color_raw": record["color_raw"],
            "order_no": record["order_no"],
            "sample_type": record["sample_type"],
            "overall_result": record["overall_result"],
            "modified_time": record["modified_time"],
            "url": url,
            "selected": selected,
            "selected_colors": "，".join(record.get("selected_colors", [])),
            "processing_status": status,
            "error_reason": reason,
        })

used_items = {row["standard_item"] for row in detail_rows}
standard_items = [item for item in ORDER if item in used_items] + sorted(item for item in used_items if item not in ORDER)
mapping_rows = sorted(mapping.values(), key=lambda row: (row["standard_item"], row["institution"], row["raw_item"]))
candidate_rows = [
    {
        "simple_item": item,
        "count": data["count"],
        "examples": data["examples"],
    }
    for item, data in sorted(item_candidates.items(), key=lambda pair: (-pair[1]["count"], pair[0]))
]
(ROOT / "item_candidates.json").write_text(json.dumps(candidate_rows, ensure_ascii=False, indent=2), encoding="utf-8")

payload = {
    "summary_rows": summary_rows,
    "detail_rows": detail_rows,
    "mapping_rows": mapping_rows,
    "source_index": source_index,
    "errors": final_errors,
    "standard_items": standard_items,
    "item_candidates": candidate_rows,
    "stats": {
        "skus": len(source["selected_skus"]),
        "summary_rows": len(summary_rows),
        "pdfs": len(source["selected_urls"]),
        "download_success": sum(document["status"] != "failed" for document in documents),
        "parsed_pdfs": sum(bool(parsed["items"]) for parsed in parsed_documents.values()),
        "detail_rows": len(detail_rows),
        "source_records": len(source["records"]),
        "invalid_records": len(source.get("invalid_records", [])),
        "errors": len(final_errors),
        "standard_items": len(standard_items),
        "item_candidates": len(candidate_rows),
    },
}
(ROOT / "trial_data.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(payload["stats"], ensure_ascii=False, indent=2))
print("error_types:", dict(Counter(row["type"] for row in final_errors)))
print("processing_status:", dict(Counter(row["processing_status"] for row in source_index)))
